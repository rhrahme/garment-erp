#!/usr/bin/env python3
"""
Generate per-client fabric reports from the Mahrab pattern import.

For each matched ERP client folder, extracts fabric numbers from attachment
filenames and companion Excel cells, looks up supplier catalog specs + local
swatch images, and writes print-friendly HTML (+ Chrome PDF when available).

Usage:
  python3 scripts/generate-mahrab-client-fabric-reports.py
  python3 scripts/generate-mahrab-client-fabric-reports.py --limit 3
  python3 scripts/generate-mahrab-client-fabric-reports.py --html-only
  python3 scripts/generate-mahrab-client-fabric-reports.py \\
    --log /tmp/mahrab-pattern-import-log.json \\
    --root "/Users/ralphrahme/Downloads/Mahrab pattern" \\
    --out "/Users/ralphrahme/Downloads/Mahrab-pattern-reports"
"""

from __future__ import annotations

import argparse
import html
import json
import mimetypes
import re
import shutil
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

try:
    from PIL import Image
except ImportError:  # pragma: no cover
    Image = None  # type: ignore

REPO = Path(__file__).resolve().parents[1]
DEFAULT_LOG = Path("/tmp/mahrab-pattern-import-log.json")
DEFAULT_ROOT = Path("/Users/ralphrahme/Downloads/Mahrab pattern")
DEFAULT_OUT = Path("/Users/ralphrahme/Downloads/Mahrab-pattern-reports")
CLIENTS_PATH = REPO / "src/data/clients.json"
PATTERN_LIBRARY_PATH = REPO / "src/data/pattern-library.json"
SUPPLIER_JSON_DIR = REPO / "src/data/suppliers"
IMAGE_ROOTS = {
    "loro-piana": REPO / "data/suppliers/loro-piana/images",
    "solbiati": REPO / "data/suppliers/loro-piana/images",
    "caccioppoli": REPO / "data/suppliers/caccioppoli/images",
    "drapers": REPO / "data/suppliers/drapers/images",
}

# Prefer full price-list catalogs over swatch-only indexes.
CATALOG_FILES = [
    "loro-piana-ss26.json",
    "caccioppoli-jackets-ss26.json",
    "caccioppoli-shirting-ss26.json",
    "drapers-hs-ss26.json",
    "zegna-ss26.json",
    "stylbiella-ss26.json",
    "stylbiella-ss25.json",
    "stylbiella-aw25.json",
    "canclini-linen-stock.json",
    "wool-stock.json",
    "gazaba-cutlength-price-list.json",
    "drapers-swatch-index.json",  # fallback for image filenames only
]

SUPPLIER_LABELS = {
    "loro-piana": "Loro Piana",
    "solbiati": "Solbiati",
    "caccioppoli": "Caccioppoli",
    "drapers": "Drapers",
    "zegna": "Zegna",
    "stylbiella": "Stylbiella",
    "canclini": "Canclini",
    "wool-stock": "Wool Stock",
    "gazaba": "Gazaba",
}

GARMENT_HINTS = [
    ("overcoat", "overcoat"),
    ("over coat", "overcoat"),
    ("overshirt", "shirt"),
    ("over shirt", "shirt"),
    ("o-s trouser", "trouser"),
    ("payjama", "custom"),
    ("pajama", "custom"),
    ("hoodie", "custom"),
    ("hoddie", "custom"),
    ("boxer", "custom"),
    ("shorts", "shorts"),
    ("short", "shorts"),
    ("trouser", "trouser"),
    ("pant", "trouser"),
    ("jacket", "jacket"),
    ("blazer", "jacket"),
    ("shirt", "shirt"),
    ("thobe", "thobe"),
    ("vest", "vest"),
    ("coat", "overcoat"),
]

SKIP_DIR_NAMES = {
    "copy",
    "marker",
    "new folder",
    "3d clo",
    ".ds_store",
}

DATE_IN_NAME_RE = re.compile(
    r"(?<!\d)(?:\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4}|\d{4}[.\-/]\d{1,2}[.\-/]\d{1,2})(?!\d)"
)
YEAR_TOKEN_RE = re.compile(r"^20[2-3]\d$")

# Strong fabric tokens from filenames / cells.
FABRIC_TOKEN_RES = [
    re.compile(r"\bDP[- ]?(\d{4,6})\b", re.I),
    re.compile(r"\bNS\d{5,6}\b", re.I),
    re.compile(r"\bN\d{5,6}\b", re.I),
    re.compile(r"\bS\d{5,6}\b", re.I),
    re.compile(r"\b\d{5}/\d{2,4}\b"),
    re.compile(r"\b\d{5}-\d{2,3}\b"),
    re.compile(r"\b\d{6}\b"),
    re.compile(r"\b\d{5}\b"),
]

FALSE_POSITIVE_CODES = {
    "10000",
    "10015",  # common model / size noise in this corpus
}


@dataclass
class FabricHit:
    raw: str
    normalized: str
    sources: set[str] = field(default_factory=set)
    garments: set[str] = field(default_factory=set)
    files: set[str] = field(default_factory=set)


@dataclass
class CatalogHit:
    fabric_number: str
    supplier_id: str
    supplier_name: str
    composition: str | None = None
    weight_gsm: Any = None
    width_cm: Any = None
    description: str | None = None
    collection: str | None = None
    color: str | None = None
    book_number: str | None = None
    mill_name: str | None = None
    swatch_filename: str | None = None
    catalog_file: str | None = None


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def client_display_name(client: dict[str, Any]) -> str:
    parts = [
        str(client.get("first_name") or "").strip(),
        str(client.get("middle_name") or "").strip(),
        str(client.get("last_name") or "").strip(),
    ]
    return " ".join(p for p in parts if p) or str(client.get("code") or "Unknown")


def load_clients() -> dict[str, dict[str, Any]]:
    data = json.loads(CLIENTS_PATH.read_text(encoding="utf-8"))
    out: dict[str, dict[str, Any]] = {}
    for c in data.get("clients") or []:
        code = str(c.get("code") or "").strip()
        if code:
            out[code] = c
    return out


def detect_garment(*texts: str) -> str:
    blob = " ".join(texts).lower()
    for hint, garment in GARMENT_HINTS:
        if hint in blob:
            return garment
    return "custom"


def strip_dates(text: str) -> str:
    return DATE_IN_NAME_RE.sub(" ", text)


def normalize_solbiati(code: str) -> str:
    upper = code.strip().upper()
    if not re.fullmatch(r"S\d+", upper):
        return upper
    typo = re.fullmatch(r"S(\d{2})0(\d{3})", upper)
    if typo:
        return f"S{typo.group(1)}{typo.group(2)}"
    return upper


def expand_candidates(raw: str) -> list[str]:
    trimmed = str(raw).strip()
    if not trimmed:
        return []
    upper = trimmed.upper().replace(" ", "")
    out: set[str] = {trimmed, upper}

    dp = re.fullmatch(r"DP[-]?(\d{4,6})", upper)
    if dp:
        out.add(dp.group(1))

    if upper.startswith("NS") and upper[2:].isdigit():
        out.add("S" + upper[2:])
        out.add(normalize_solbiati("S" + upper[2:]))
        out.add(upper[2:])

    if re.fullmatch(r"N\d+", upper):
        out.add(upper[1:])

    if re.fullmatch(r"S\d+", upper):
        out.add(normalize_solbiati(upper))

    if re.fullmatch(r"\d{5}-\d{2,3}", upper):
        left, right = upper.split("-", 1)
        out.add(left)
        out.add(f"{left}/{right}")

    if re.fullmatch(r"\d{5}/\d{2,4}", upper):
        out.add(upper.split("/", 1)[0])
        out.add(upper.replace("/", "-"))

    if re.fullmatch(r"\d{5}", upper):
        out.add("S" + upper)

    if re.fullmatch(r"\d{6}", upper):
        out.add(upper)

    # Stable order: prefer catalog-looking forms first
    preferred = sorted(
        out,
        key=lambda c: (
            0 if re.fullmatch(r"S\d{5}", c, re.I) else 1,
            0 if re.fullmatch(r"\d{6}", c) else 1,
            0 if "/" in c else 1,
            len(c),
            c,
        ),
    )
    return preferred


def preferred_display_code(raw: str, catalog_number: str | None = None) -> str:
    if catalog_number:
        return catalog_number
    upper = str(raw).strip().upper().replace(" ", "")
    # Keep explicit Solbiati / LP / DP / stylbiella forms as written.
    if upper.startswith("NS") and upper[2:].isdigit():
        return normalize_solbiati("S" + upper[2:])
    if re.fullmatch(r"S\d+", upper):
        return normalize_solbiati(upper)
    if re.fullmatch(r"N\d+", upper):
        return upper[1:]
    dp = re.fullmatch(r"DP[-]?(\d{4,6})", upper)
    if dp:
        return dp.group(1)
    if re.fullmatch(r"\d{5}-\d{2,3}", upper):
        left, right = upper.split("-", 1)
        return f"{left}/{right}"
    if re.fullmatch(r"\d{5}/\d{2,4}", upper):
        return upper
    if re.fullmatch(r"\d{6}", upper) or re.fullmatch(r"\d{5}", upper):
        return upper
    cands = expand_candidates(raw)
    return (cands[0] if cands else raw).upper()


def is_plausible_fabric_token(token: str) -> bool:
    t = token.strip().upper()
    if not t or t in FALSE_POSITIVE_CODES:
        return False
    if YEAR_TOKEN_RE.fullmatch(t):
        return False
    if re.fullmatch(r"\d{1,2}", t):
        return False
    # Reject bare day/month-like leftovers
    if re.fullmatch(r"0?\d{1,2}", t):
        return False
    if re.fullmatch(r"DP-?\d{4,6}", t):
        return True
    if re.fullmatch(r"NS\d{5,6}", t):
        return True
    if re.fullmatch(r"N\d{5,6}", t):
        return True
    if re.fullmatch(r"S\d{5,6}", t):
        return True
    if re.fullmatch(r"\d{5}/\d{2,4}", t):
        return True
    if re.fullmatch(r"\d{5}-\d{2,3}", t):
        return True
    if re.fullmatch(r"\d{6}", t):
        return True
    if re.fullmatch(r"\d{5}", t):
        # 5-digit alone is weak; keep only when later catalog-confirmed or DP/NS context
        return True
    return False


def extract_tokens_from_text(text: str) -> list[str]:
    cleaned = strip_dates(text)
    found: list[str] = []
    seen: set[str] = set()
    for cre in FABRIC_TOKEN_RES:
        for m in cre.finditer(cleaned):
            token = m.group(0)
            if cre.pattern.startswith(r"\bDP"):
                token = "DP-" + m.group(1)
            key = token.upper()
            if key in seen:
                continue
            if not is_plausible_fabric_token(token):
                continue
            seen.add(key)
            found.append(token)
    return found


def walk_files(root: Path) -> list[Path]:
    out: list[Path] = []
    if not root.exists():
        return out
    stack = [root]
    while stack:
        cur = stack.pop()
        try:
            entries = list(cur.iterdir())
        except OSError:
            continue
        for ent in entries:
            name = ent.name
            if name.startswith(".") or name.startswith("~$"):
                continue
            if ent.is_dir():
                if name.lower() in SKIP_DIR_NAMES:
                    continue
                stack.append(ent)
            else:
                out.append(ent)
    return out


def extract_from_xlsx(path: Path) -> list[str]:
    if load_workbook is None:
        return []
    tokens: list[str] = []
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return []
    try:
        for ws in wb.worksheets:
            rows: list[list[Any]] = []
            for i, row in enumerate(ws.iter_rows(max_row=80, max_col=20, values_only=True)):
                rows.append(list(row))
                if i >= 79:
                    break
            # Fabric label neighbour
            for r_idx, row in enumerate(rows):
                for c_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    label = str(cell).strip().lower()
                    if label in {"fabric", "fabric:", "fabric code", "fabric name/code", "fabric name/code :"} or (
                        "fabric" in label and ("code" in label or "name" in label)
                    ):
                        # next cells on same row
                        for nxt in row[c_idx + 1 : c_idx + 4]:
                            if nxt is None or str(nxt).strip() == "":
                                continue
                            tokens.extend(extract_tokens_from_text(str(nxt)))
                            # bare numeric fabric in Fabric column
                            if isinstance(nxt, (int, float)) and not isinstance(nxt, bool):
                                num = str(int(nxt)) if float(nxt).is_integer() else str(nxt)
                                tokens.extend(extract_tokens_from_text(num))
                            break
            # any cell that looks like a fabric token
            for row in rows:
                for cell in row:
                    if cell is None:
                        continue
                    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
                        num = str(int(cell)) if float(cell).is_integer() else str(cell)
                        if re.fullmatch(r"\d{5,6}", num):
                            tokens.append(num)
                    else:
                        tokens.extend(extract_tokens_from_text(str(cell)))
    finally:
        wb.close()
    # de-dupe preserve order
    out: list[str] = []
    seen: set[str] = set()
    for t in tokens:
        k = t.upper()
        if k in seen:
            continue
        seen.add(k)
        out.append(t)
    return out


def supplier_id_from_stem(stem: str, fabric_number: str) -> str:
    if "loro-piana" in stem:
        return "solbiati" if re.match(r"^S\d", fabric_number, re.I) else "loro-piana"
    if "caccioppoli" in stem:
        return "caccioppoli"
    if "drapers" in stem:
        return "drapers"
    if "zegna" in stem:
        return "zegna"
    if "stylbiella" in stem:
        return "stylbiella"
    if "canclini" in stem:
        return "canclini"
    if "wool-stock" in stem:
        return "wool-stock"
    if "gazaba" in stem:
        return "gazaba"
    return stem


def load_catalog_index() -> dict[str, CatalogHit]:
    """Map uppercased fabric_number -> best CatalogHit."""
    index: dict[str, CatalogHit] = {}
    for filename in CATALOG_FILES:
        path = SUPPLIER_JSON_DIR / filename
        if not path.exists():
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        fabrics = data.get("fabrics") if isinstance(data, dict) else None
        if not isinstance(fabrics, list):
            continue
        for fab in fabrics:
            num = str(fab.get("fabric_number") or "").strip()
            if not num:
                continue
            sid = supplier_id_from_stem(path.stem, num)
            hit = CatalogHit(
                fabric_number=num,
                supplier_id=sid,
                supplier_name=SUPPLIER_LABELS.get(sid, sid),
                composition=fab.get("composition"),
                weight_gsm=fab.get("weight_gsm") or fab.get("weight_linear"),
                width_cm=fab.get("width_cm"),
                description=fab.get("description"),
                collection=fab.get("collection"),
                color=fab.get("color"),
                book_number=fab.get("book_number"),
                mill_name=fab.get("mill_name") or fab.get("mill_code"),
                swatch_filename=fab.get("swatch_filename"),
                catalog_file=filename,
            )
            key = num.upper()
            existing = index.get(key)
            if existing is None:
                index[key] = hit
                continue
            # Prefer richer specs over swatch-index stubs
            existing_score = sum(
                1
                for v in (
                    existing.composition,
                    existing.weight_gsm,
                    existing.width_cm,
                    existing.description,
                )
                if v not in (None, "")
            )
            new_score = sum(
                1
                for v in (hit.composition, hit.weight_gsm, hit.width_cm, hit.description)
                if v not in (None, "")
            )
            if new_score > existing_score:
                index[key] = hit
    return index


def lookup_catalog(raw: str, index: dict[str, CatalogHit]) -> CatalogHit | None:
    for cand in expand_candidates(raw):
        hit = index.get(cand.upper())
        if hit:
            return hit
    return None


def find_swatch_path(raw: str, catalog: CatalogHit | None) -> Path | None:
    candidates: list[str] = []
    if catalog and catalog.swatch_filename:
        candidates.append(catalog.swatch_filename)
    for cand in expand_candidates(raw):
        candidates.append(cand)
        if catalog:
            candidates.append(catalog.fabric_number)
    # unique preserve
    seen: set[str] = set()
    uniq: list[str] = []
    for c in candidates:
        k = c.upper()
        if k in seen:
            continue
        seen.add(k)
        uniq.append(c)

    roots: list[Path] = []
    if catalog:
        root = IMAGE_ROOTS.get(catalog.supplier_id)
        if root:
            roots.append(root)
    roots.extend([IMAGE_ROOTS["loro-piana"], IMAGE_ROOTS["caccioppoli"], IMAGE_ROOTS["drapers"]])
    # unique roots
    root_seen: set[Path] = set()
    ordered_roots: list[Path] = []
    for r in roots:
        if r in root_seen:
            continue
        root_seen.add(r)
        ordered_roots.append(r)

    for stem in uniq:
        base = Path(stem).name
        stem_only = Path(base).stem if "." in base else base
        for root in ordered_roots:
            if not root.exists():
                continue
            # exact filename from catalog
            if "." in base:
                p = root / base
                if p.exists():
                    return p
            for ext in (".jpg", ".jpeg", ".png", ".webp"):
                p = root / f"{stem_only}{ext}"
                if p.exists():
                    return p
                p2 = root / f"{stem_only.upper()}{ext}"
                if p2.exists():
                    return p2
                p3 = root / f"{stem_only.lower()}{ext}"
                if p3.exists():
                    return p3
    return None


def cache_swatch_asset(src: Path, assets_dir: Path, fabric_key: str) -> str | None:
    """Copy/resize swatch into assets_dir; return relative path assets/<name>."""
    assets_dir.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", fabric_key.upper())[:80] or "swatch"
    dest_name = f"{safe}.jpg"
    dest = assets_dir / dest_name
    if not dest.exists():
        try:
            if Image is not None:
                with Image.open(src) as im:
                    im = im.convert("RGB")
                    im.thumbnail((320, 320))
                    im.save(dest, format="JPEG", quality=82, optimize=True)
            else:
                shutil.copy2(src, dest)
        except OSError:
            return None
    return f"assets/{dest_name}"


def collect_client_hits(
    folder: Path,
    client_code: str,
    pattern_library: dict[str, Any],
) -> dict[str, FabricHit]:
    hits: dict[str, FabricHit] = {}

    def add(token: str, source: str, garment: str, file_label: str) -> None:
        if not is_plausible_fabric_token(token):
            return
        key = preferred_display_code(token).upper()
        # Prefer a stable key using first candidate
        cands = expand_candidates(token)
        key = (cands[0] if cands else token).upper()
        hit = hits.get(key)
        if hit is None:
            hit = FabricHit(raw=token, normalized=preferred_display_code(token), sources=set(), garments=set(), files=set())
            hits[key] = hit
        hit.sources.add(source)
        hit.garments.add(garment)
        hit.files.add(file_label)
        # Prefer richer raw forms
        if re.match(r"^(NS|N|S|DP)", token, re.I) and not re.match(r"^(NS|N|S|DP)", hit.raw, re.I):
            hit.raw = token

    if folder.exists():
        for path in walk_files(folder):
            rel = str(path.relative_to(folder))
            garment = detect_garment(rel, path.name)
            ext = path.suffix.lower()
            if ext in {".xlsx", ".xls", ".tud", ".pdf", ".jpg", ".jpeg", ".png"}:
                for tok in extract_tokens_from_text(path.name):
                    add(tok, "filename", garment, rel)
            if ext in {".xlsx", ".xls"}:
                for tok in extract_from_xlsx(path):
                    add(tok, "xlsx_cell", garment, rel)

    # Pattern library attachments / file names for this client
    for cp in pattern_library.get("client_patterns") or []:
        if str(cp.get("client_code") or "") != client_code:
            continue
        garment = str(cp.get("garment_type") or "custom")
        for f in cp.get("files") or []:
            fname = str(f.get("filename") or "")
            if not fname:
                continue
            for tok in extract_tokens_from_text(fname):
                add(tok, "pattern_library", garment, fname)
        for ver in cp.get("versions") or []:
            for f in ver.get("files") or []:
                fname = str(f.get("filename") or "")
                if not fname:
                    continue
                for tok in extract_tokens_from_text(fname):
                    add(tok, "pattern_library", garment, fname)
        for line_id in cp.get("linked_fabric_line_ids") or []:
            # keep as source marker only; line ids are not fabric numbers
            _ = line_id

    return hits


def refine_hits_with_catalog(
    hits: dict[str, FabricHit], index: dict[str, CatalogHit]
) -> list[tuple[FabricHit, CatalogHit | None, Path | None]]:
    """Drop weak 5-digit orphans that never match a catalog; keep strong tokens."""
    rows: list[tuple[FabricHit, CatalogHit | None, Path | None]] = []
    for hit in hits.values():
        catalog = lookup_catalog(hit.raw, index) or lookup_catalog(hit.normalized, index)
        if catalog:
            hit.normalized = catalog.fabric_number
        swatch = find_swatch_path(hit.raw, catalog)
        weak_five = bool(re.fullmatch(r"\d{5}", hit.raw.upper()) or re.fullmatch(r"\d{5}", hit.normalized.upper()))
        strong = bool(
            re.match(r"^(NS|N|S|DP)", hit.raw, re.I)
            or re.fullmatch(r"\d{6}", hit.raw.upper())
            or "/" in hit.raw
            or re.fullmatch(r"\d{5}-\d{2,3}", hit.raw.upper())
            or "xlsx_cell" in hit.sources
        )
        if catalog is None and weak_five and not strong:
            continue
        rows.append((hit, catalog, swatch))

    # Merge duplicates that resolve to same catalog number / digit core
    def merge_key(hit: FabricHit, catalog: CatalogHit | None) -> str:
        if catalog:
            return catalog.fabric_number.upper()
        n = hit.normalized.upper()
        if n.startswith("S") and n[1:].isdigit():
            # Only keep S-form if raw was explicitly Solbiati
            if re.match(r"^(NS|S)\d", hit.raw, re.I):
                return n
            return n[1:]
        return n

    merged: dict[str, tuple[FabricHit, CatalogHit | None, Path | None]] = {}
    for hit, catalog, swatch in rows:
        # Drop phantom S-prefixed forms invented from bare digits
        if (
            catalog is None
            and re.fullmatch(r"S\d{5}", hit.normalized.upper())
            and not re.match(r"^(NS|S)\d", hit.raw, re.I)
        ):
            continue
        key = merge_key(hit, catalog)
        if key not in merged:
            merged[key] = (hit, catalog, swatch)
            continue
        prev_hit, prev_cat, prev_swatch = merged[key]
        prev_hit.sources |= hit.sources
        prev_hit.garments |= hit.garments
        prev_hit.files |= hit.files
        if catalog and not prev_cat:
            prev_hit.normalized = catalog.fabric_number
            merged[key] = (prev_hit, catalog, swatch or prev_swatch)
        elif swatch and not prev_swatch:
            merged[key] = (prev_hit, prev_cat, swatch)

    # Drop bare 5-digit codes when a more specific article/color form exists
    specifics: set[str] = set()
    for hit, catalog, _swatch in merged.values():
        num = (catalog.fabric_number if catalog else hit.normalized).upper()
        m = re.fullmatch(r"(\d{5})[/-](\d{2,4})", num)
        if m:
            specifics.add(m.group(1))
    if specifics:
        merged = {
            k: v
            for k, v in merged.items()
            if not (
                re.fullmatch(r"\d{5}", (v[1].fabric_number if v[1] else v[0].normalized).upper())
                and (v[1].fabric_number if v[1] else v[0].normalized).upper() in specifics
            )
        }

    out = list(merged.values())
    out.sort(key=lambda row: (sorted(row[0].garments)[0] if row[0].garments else "zzz", row[0].normalized))
    return out


def slugify(text: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "-", text.strip()).strip("-").lower()
    return s or "client"


def render_client_html(
    code: str,
    name: str,
    folder: str,
    rows: list[tuple[FabricHit, CatalogHit | None, Path | None]],
    generated_at: str,
    assets_dir: Path,
) -> str:
    # Group by garment
    by_garment: dict[str, list[tuple[FabricHit, CatalogHit | None, Path | None]]] = defaultdict(list)
    for row in rows:
        garments = sorted(row[0].garments) or ["unspecified"]
        for g in garments:
            by_garment[g].append(row)

    fabric_cards: list[str] = []
    # Unique fabrics section (one card per fabric)
    seen_fabrics: set[str] = set()
    for hit, catalog, swatch in rows:
        key = hit.normalized.upper()
        if key in seen_fabrics:
            continue
        seen_fabrics.add(key)

        img_html = '<div class="swatch missing">No swatch image</div>'
        if swatch:
            rel = cache_swatch_asset(swatch, assets_dir, hit.normalized)
            if rel:
                img_html = f'<img class="swatch" src="{html.escape(rel)}" alt="{html.escape(hit.normalized)}" />'

        if catalog:
            spec_bits = [
                ("Supplier", catalog.supplier_name),
                ("Catalog #", catalog.fabric_number),
                ("Composition", catalog.composition or "-"),
                ("Weight", f"{catalog.weight_gsm}" if catalog.weight_gsm not in (None, "") else "-"),
                ("Width", f"{catalog.width_cm} cm" if catalog.width_cm not in (None, "") else "-"),
                ("Collection", catalog.collection or "-"),
                ("Color", catalog.color or "-"),
                ("Mill", catalog.mill_name or "-"),
                ("Description", catalog.description or "-"),
            ]
        else:
            spec_bits = [
                ("Supplier", "Not found in catalogs"),
                ("Catalog #", "-"),
                ("Composition", "-"),
                ("Weight", "-"),
                ("Width", "-"),
                ("Collection", "-"),
                ("Color", "-"),
                ("Mill", "-"),
                ("Description", "-"),
            ]

        specs = "".join(
            f"<div class='spec'><span>{html.escape(k)}</span><strong>{html.escape(str(v))}</strong></div>"
            for k, v in spec_bits
        )
        sources = ", ".join(sorted(hit.sources)) or "-"
        garments = ", ".join(sorted(hit.garments)) or "-"
        files = "<br/>".join(html.escape(f) for f in sorted(hit.files)[:8])
        if len(hit.files) > 8:
            files += f"<br/>... +{len(hit.files) - 8} more"

        fabric_cards.append(
            f"""
            <article class="fabric-card">
              <div class="media">{img_html}</div>
              <div class="body">
                <h3>{html.escape(hit.normalized)} <small>raw: {html.escape(hit.raw)}</small></h3>
                <div class="meta">Garments: {html.escape(garments)} | Sources: {html.escape(sources)}</div>
                <div class="specs">{specs}</div>
                <div class="files"><strong>Evidence files</strong><div>{files or "-"}</div></div>
              </div>
            </article>
            """
        )

    garment_sections: list[str] = []
    for garment in sorted(by_garment.keys()):
        codes = []
        seen: set[str] = set()
        for hit, catalog, _swatch in by_garment[garment]:
            label = catalog.fabric_number if catalog else hit.normalized
            if label.upper() in seen:
                continue
            seen.add(label.upper())
            codes.append(label)
        code_list = ", ".join(html.escape(c) for c in codes) if codes else "-"
        garment_sections.append(
            f"<tr><td>{html.escape(garment)}</td><td>{code_list}</td><td>{len(codes)}</td></tr>"
        )

    if not rows:
        body_main = "<p class='empty'>No fabric codes in files.</p>"
    else:
        body_main = f"""
        <section>
          <h2>Fabrics by garment</h2>
          <table>
            <thead><tr><th>Garment</th><th>Fabric numbers</th><th>Count</th></tr></thead>
            <tbody>
              {''.join(garment_sections)}
            </tbody>
          </table>
        </section>
        <section>
          <h2>Fabric specifications ({len(seen_fabrics)})</h2>
          <div class="cards">
            {''.join(fabric_cards)}
          </div>
        </section>
        """

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>{html.escape(code)} - {html.escape(name)} fabric report</title>
  <style>
    :root {{
      --ink: #1c1917;
      --muted: #57534e;
      --line: #e7e5e4;
      --bg: #fafaf9;
      --card: #ffffff;
      --accent: #0f766e;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Iowan Old Style", "Palatino Linotype", Palatino, Georgia, serif;
      color: var(--ink);
      background: linear-gradient(180deg, #f5f5f4 0%, var(--bg) 220px);
      line-height: 1.45;
    }}
    .wrap {{ max-width: 980px; margin: 0 auto; padding: 32px 24px 64px; }}
    header {{
      border-bottom: 2px solid var(--ink);
      padding-bottom: 16px;
      margin-bottom: 28px;
    }}
    .brand {{
      font-size: 12px;
      letter-spacing: 0.14em;
      text-transform: uppercase;
      color: var(--accent);
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      margin: 0 0 8px;
    }}
    h1 {{
      font-size: 28px;
      margin: 0 0 6px;
      font-weight: 600;
    }}
    .sub {{
      color: var(--muted);
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      font-size: 13px;
    }}
    h2 {{
      font-size: 18px;
      margin: 28px 0 12px;
      border-bottom: 1px solid var(--line);
      padding-bottom: 6px;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      font-size: 13px;
      background: var(--card);
    }}
    th, td {{
      border: 1px solid var(--line);
      padding: 8px 10px;
      text-align: left;
      vertical-align: top;
    }}
    th {{ background: #f5f5f4; }}
    .cards {{ display: grid; gap: 16px; }}
    .fabric-card {{
      display: grid;
      grid-template-columns: 160px 1fr;
      gap: 16px;
      background: var(--card);
      border: 1px solid var(--line);
      padding: 14px;
      break-inside: avoid;
      page-break-inside: avoid;
    }}
    .swatch {{
      width: 160px;
      height: 160px;
      object-fit: cover;
      background: #e7e5e4;
      display: block;
    }}
    .swatch.missing {{
      display: flex;
      align-items: center;
      justify-content: center;
      color: var(--muted);
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      font-size: 12px;
      text-align: center;
      padding: 12px;
    }}
    h3 {{
      margin: 0 0 4px;
      font-size: 18px;
    }}
    h3 small {{
      font-weight: 400;
      color: var(--muted);
      font-size: 12px;
      font-family: "Avenir Next", "Segoe UI", sans-serif;
    }}
    .meta, .files {{
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      font-size: 12px;
      color: var(--muted);
      margin-bottom: 10px;
    }}
    .specs {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 6px 14px;
      margin-bottom: 10px;
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      font-size: 12px;
    }}
    .spec span {{
      display: block;
      color: var(--muted);
      font-size: 10px;
      text-transform: uppercase;
      letter-spacing: 0.06em;
    }}
    .spec strong {{ font-weight: 600; color: var(--ink); }}
    .empty {{
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      color: var(--muted);
      padding: 24px;
      border: 1px dashed var(--line);
      background: var(--card);
    }}
    footer {{
      margin-top: 36px;
      padding-top: 12px;
      border-top: 1px solid var(--line);
      font-family: "Avenir Next", "Segoe UI", sans-serif;
      font-size: 11px;
      color: var(--muted);
    }}
    @media print {{
      body {{ background: white; }}
      .wrap {{ max-width: none; padding: 12mm; }}
      .fabric-card {{ break-inside: avoid; }}
    }}
    @media (max-width: 720px) {{
      .fabric-card {{ grid-template-columns: 1fr; }}
      .swatch, .swatch.missing {{ width: 100%; height: 200px; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <header>
      <p class="brand">Hagan | Mahrab pattern import</p>
      <h1>{html.escape(name)}</h1>
      <p class="sub">Client code {html.escape(code)} | Folder "{html.escape(folder)}" | Generated {html.escape(generated_at)}</p>
    </header>
    {body_main}
    <footer>
      Source: Mahrab pattern folder + supplier catalogs (Loro Piana / Solbiati, Caccioppoli, Drapers, Zegna, Stylbiella, ...).
      Swatches from local caches under data/suppliers/*/images. Regenerate via scripts/generate-mahrab-client-fabric-reports.py.
    </footer>
  </div>
</body>
</html>
"""


def render_index_html(
    clients: list[dict[str, Any]],
    generated_at: str,
    out_dir: Path,
) -> str:
    rows = []
    for c in clients:
        href = html.escape(c["html_name"])
        pdf_cell = "-"
        if c.get("pdf_name"):
            pdf_cell = f'<a href="{html.escape(c["pdf_name"])}">PDF</a>'
        rows.append(
            f"<tr>"
            f"<td>{html.escape(c['code'])}</td>"
            f"<td><a href='{href}'>{html.escape(c['name'])}</a></td>"
            f"<td>{c['fabric_count']}</td>"
            f"<td>{c['with_spec']}</td>"
            f"<td>{c['with_image']}</td>"
            f"<td>{pdf_cell}</td>"
            f"</tr>"
        )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <title>Mahrab pattern - client fabric reports</title>
  <style>
    body {{ font-family: "Avenir Next", "Segoe UI", sans-serif; margin: 32px; color: #1c1917; }}
    h1 {{ font-family: Georgia, serif; }}
    table {{ border-collapse: collapse; width: 100%; max-width: 960px; }}
    th, td {{ border: 1px solid #e7e5e4; padding: 8px 10px; text-align: left; font-size: 13px; }}
    th {{ background: #f5f5f4; }}
    .muted {{ color: #57534e; font-size: 13px; }}
  </style>
</head>
<body>
  <h1>Mahrab pattern - client fabric reports</h1>
  <p class="muted">Generated {html.escape(generated_at)} | {len(clients)} clients | output {html.escape(str(out_dir))}</p>
  <table>
    <thead>
      <tr><th>Code</th><th>Client</th><th>Fabrics</th><th>With spec</th><th>With image</th><th>PDF</th></tr>
    </thead>
    <tbody>
      {''.join(rows)}
    </tbody>
  </table>
</body>
</html>
"""


def chrome_print_pdf(html_path: Path, pdf_path: Path) -> bool:
    chrome = Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome")
    if not chrome.exists():
        return False
    cmd = [
        str(chrome),
        "--headless=new",
        "--disable-gpu",
        "--allow-file-access-from-files",
        "--no-pdf-header-footer",
        f"--print-to-pdf={pdf_path}",
        html_path.resolve().as_uri(),
    ]
    try:
        subprocess.run(cmd, check=True, capture_output=True, timeout=120)
        return pdf_path.exists() and pdf_path.stat().st_size > 0
    except (subprocess.SubprocessError, OSError):
        return False


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--log", type=Path, default=DEFAULT_LOG)
    parser.add_argument("--root", type=Path, default=DEFAULT_ROOT)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--limit", type=int, default=0, help="Process only first N matched clients")
    parser.add_argument("--html-only", action="store_true", help="Skip Chrome PDF generation")
    parser.add_argument(
        "--skip-empty",
        action="store_true",
        help="Skip clients with zero fabric codes (default: include with notice)",
    )
    args = parser.parse_args(argv)

    if not args.log.exists():
        print(f"ERROR: import log not found: {args.log}", file=sys.stderr)
        return 1
    if not args.root.exists():
        print(f"ERROR: Mahrab root not found: {args.root}", file=sys.stderr)
        return 1

    log = json.loads(args.log.read_text(encoding="utf-8"))
    matched = log.get("matched_folders") or []
    if not matched:
        print("ERROR: no matched_folders in import log", file=sys.stderr)
        return 1

    clients_by_code = load_clients()
    pattern_library = (
        json.loads(PATTERN_LIBRARY_PATH.read_text(encoding="utf-8"))
        if PATTERN_LIBRARY_PATH.exists()
        else {"client_patterns": []}
    )
    catalog_index = load_catalog_index()
    print(f"Catalog index: {len(catalog_index)} fabric numbers")

    args.out.mkdir(parents=True, exist_ok=True)
    generated_at = utc_now()

    # Stable unique by code (keep first folder match)
    by_code: dict[str, dict[str, Any]] = {}
    for row in matched:
        code = str(row.get("code") or "").strip()
        folder = str(row.get("folder") or "").strip()
        if not code or not folder:
            continue
        if code not in by_code:
            by_code[code] = row

    codes = sorted(by_code.keys())
    if args.limit and args.limit > 0:
        codes = codes[: args.limit]

    index_rows: list[dict[str, Any]] = []
    link_payload: dict[str, Any] = {
        "generated_at": generated_at,
        "clients": [],
    }
    summary = {
        "generated_at": generated_at,
        "clients": 0,
        "with_fabrics": 0,
        "empty": 0,
        "pdfs": 0,
        "items": [],
    }

    for code in codes:
        row = by_code[code]
        folder_name = str(row["folder"])
        client = clients_by_code.get(code) or {}
        name = client_display_name(client) if client else folder_name
        folder_path = args.root / folder_name

        hits = collect_client_hits(folder_path, code, pattern_library)
        refined = refine_hits_with_catalog(hits, catalog_index)

        if not refined and args.skip_empty:
            summary["empty"] += 1
            print(f"SKIP empty {code} {name}")
            continue

        with_spec = sum(1 for _h, c, _s in refined if c is not None)
        with_image = sum(1 for _h, _c, s in refined if s is not None)

        html_name = f"{code}-{slugify(name)}.html"
        pdf_name = f"{code}-{slugify(name)}.pdf"
        html_path = args.out / html_name
        pdf_path = args.out / pdf_name

        assets_dir = args.out / "assets"
        html_path.write_text(
            render_client_html(code, name, folder_name, refined, generated_at, assets_dir),
            encoding="utf-8",
        )

        pdf_ok = False
        if not args.html_only:
            pdf_ok = chrome_print_pdf(html_path, pdf_path)
            if pdf_ok:
                summary["pdfs"] += 1

        summary["clients"] += 1
        if refined:
            summary["with_fabrics"] += 1
        else:
            summary["empty"] += 1

        item = {
            "code": code,
            "name": name,
            "folder": folder_name,
            "fabric_count": len(refined),
            "with_spec": with_spec,
            "with_image": with_image,
            "html": html_name,
            "pdf": pdf_name if pdf_ok or pdf_path.exists() else None,
        }
        summary["items"].append(item)
        index_rows.append(
            {
                "code": code,
                "name": name,
                "fabric_count": len(refined),
                "with_spec": with_spec,
                "with_image": with_image,
                "html_name": html_name,
                "pdf_name": pdf_name if pdf_ok or pdf_path.exists() else None,
            }
        )

        link_fabrics: list[dict[str, Any]] = []
        for hit, catalog, swatch in refined:
            link_fabrics.append(
                {
                    "fabric_number": catalog.fabric_number if catalog else hit.normalized,
                    "raw": hit.raw,
                    "supplier_id": catalog.supplier_id if catalog else None,
                    "supplier_name": catalog.supplier_name if catalog else None,
                    "composition": catalog.composition if catalog else None,
                    "weight_gsm": catalog.weight_gsm if catalog else None,
                    "width_cm": catalog.width_cm if catalog else None,
                    "color": catalog.color if catalog else None,
                    "description": catalog.description if catalog else None,
                    "garments": sorted(hit.garments) or ["custom"],
                    "has_image": swatch is not None,
                    "sources": sorted(hit.sources),
                }
            )
        link_payload["clients"].append(
            {
                "code": code,
                "name": name,
                "folder": folder_name,
                "client_id": str(client.get("id") or "") or None,
                "pdf": pdf_name if pdf_ok or pdf_path.exists() else None,
                "fabrics": link_fabrics,
            }
        )
        print(
            f"OK {code} {name}: {len(refined)} fabrics "
            f"({with_spec} specs, {with_image} images) -> {html_name}"
            + (f" + {pdf_name}" if pdf_ok else " (html only)")
        )

    index_path = args.out / "index.html"
    index_path.write_text(render_index_html(index_rows, generated_at, args.out), encoding="utf-8")
    summary_path = args.out / "summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    link_path = args.out / "fabrics-link.json"
    link_path.write_text(json.dumps(link_payload, indent=2), encoding="utf-8")

    print()
    print(f"Wrote {summary['clients']} client reports to {args.out}")
    print(f"  with fabrics: {summary['with_fabrics']}  empty: {summary['empty']}  pdfs: {summary['pdfs']}")
    print(f"  index: {index_path}")
    print(f"  summary: {summary_path}")
    print(f"  fabrics-link: {link_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
