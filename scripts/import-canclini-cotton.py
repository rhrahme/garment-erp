#!/usr/bin/env python3
"""Merge Canclini cotton shirting stock into Canclini warehouse catalog JSON."""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
CANCLINI_JSON = ROOT / "src" / "data" / "suppliers" / "canclini-linen-stock.json"
LINEN_XLSX = Path.home() / "Desktop" / "Fabrics" / "Linen Stock HAGAN.xlsx"
# Source files live under ~/Desktop/Fabrics/Luthai/ (legacy folder name on disk)
COTTON_FILES = [
    Path.home() / "Desktop" / "Fabrics" / "Luthai" / "Luth PL& Inv.xlsx",
    Path.home() / "Desktop" / "Fabrics" / "Luthai" / "Luth PL& iNV 2.xlsx",
]


def normalize_fabric_number(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower().startswith("description"):
        return None
    if text.count("-") >= 2:
        return text.split("-", 1)[1].strip()
    return text


def parse_weight(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[\d.]+", str(value))
    return float(match.group(0)) if match else None


def parse_width(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"(\d+)", str(value))
    return int(match.group(1)) if match else None


def is_valid_linen_code(code: str | None) -> bool:
    if not code:
        return False
    text = str(code).strip()
    if not text or text.lower().startswith("total"):
        return False
    return bool(re.match(r"^\d{2}[TH]\d{3}$", text, re.IGNORECASE))


def parse_linen(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Codes"]
    fabrics: list[dict] = []
    seen: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        pairs = [
            (row[0], row[1], row[2], row[3], row[4]),
            (row[5], row[6], row[7], row[8], row[9]),
        ]
        for code, composition, weight, width, avail_meters in pairs:
            if not is_valid_linen_code(code):
                continue
            fabric_number = str(code).strip().upper()
            if fabric_number in seen:
                continue
            seen.add(fabric_number)

            weight_gsm = parse_weight(weight)
            width_cm = parse_width(width)
            category = "H-weight" if "H" in fabric_number[2:4] else "T-weight"
            weight_label = f"{int(weight_gsm)} g/m²" if weight_gsm else "unknown weight"

            fabrics.append(
                {
                    "fabric_number": fabric_number,
                    "composition": str(composition).strip() if composition else "100% Linen",
                    "color": None,
                    "description": f"Canclini linen — {category} ({weight_label})",
                    "weight_gsm": weight_gsm,
                    "width_cm": width_cm,
                    "collection": "HAGAN Linen Stock",
                    "category": category,
                    "unit_price": None,
                    "unit": "meters",
                    "currency": "EUR",
                    "is_active": True,
                    "stock_status": "in_stock",
                    "available_meters": float(avail_meters) if isinstance(avail_meters, (int, float)) else None,
                    "source": "linen",
                }
            )

    wb.close()
    fabrics.sort(key=lambda item: item["fabric_number"])
    return fabrics


def packing_list_sheet(wb) -> str:
    return next((name for name in wb.sheetnames if "pack" in name.lower()), wb.sheetnames[-1])


def invoice_sheet(wb) -> str:
    return next((name for name in wb.sheetnames if "invoice" in name.lower()), wb.sheetnames[0])


def parse_packing_list(xlsx_path: Path) -> dict[str, dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[packing_list_sheet(wb)]
    items: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row:
            continue
        pattern = row[1] if len(row) > 1 else None
        fabric_no = row[2] if len(row) > 2 else None
        construction = row[4] if len(row) > 4 else None
        qty = row[5] if len(row) > 5 else None
        if not pattern and not fabric_no:
            continue

        key = normalize_fabric_number(fabric_no) if fabric_no and str(fabric_no).strip() else normalize_fabric_number(pattern)
        if not key or not is_valid_cotton_key(key, str(pattern).strip() if pattern else None):
            continue

        rec = items.setdefault(
            key,
            {
                "fabric_number": key,
                "pattern_no": None,
                "fabric_raw": None,
                "construction": None,
                "available_meters": None,
                "composition": None,
                "unit_price": None,
            },
        )
        if pattern:
            rec["pattern_no"] = str(pattern).strip()
        if fabric_no:
            rec["fabric_raw"] = normalize_fabric_number(fabric_no) or str(fabric_no).strip()
        if construction:
            rec["construction"] = str(construction).strip()
        if isinstance(qty, (int, float)):
            rec["available_meters"] = (rec["available_meters"] or 0) + float(qty)

    wb.close()
    return items


def parse_invoice(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[invoice_sheet(wb)]
    rows: list[dict] = []

    for row in ws.iter_rows(min_row=13, values_only=True):
        if not row:
            continue
        desc = row[1]
        if desc is None or str(desc).strip().lower().startswith("description"):
            continue

        composition = None
        price = None
        if len(row) >= 8 and isinstance(row[3], (int, float)):
            composition = row[2]
            price = row[3]
        elif len(row) >= 3 and isinstance(row[2], (int, float)):
            price = row[2]

        pattern = str(desc).strip()
        if not is_valid_cotton_key(normalize_fabric_number(pattern), pattern):
            continue
        unit_price = float(price) if isinstance(price, (int, float)) and price > 0 else None
        comp = str(composition).strip() if composition and not isinstance(composition, (int, float)) else None

        rows.append(
            {
                "pattern_no": pattern,
                "fabric_number": normalize_fabric_number(pattern),
                "composition": comp,
                "unit_price": unit_price,
            }
        )

    wb.close()
    return rows


def apply_invoice_prices(items: dict[str, dict], invoices: list[dict]) -> None:
    by_pattern = {rec["pattern_no"]: rec for rec in items.values() if rec.get("pattern_no")}

    for inv in invoices:
        pattern = inv["pattern_no"]
        target = by_pattern.get(pattern)
        if not target and inv.get("fabric_number"):
            target = items.get(inv["fabric_number"])
        if not target:
            key = inv.get("fabric_number") or normalize_fabric_number(pattern)
            if not key or not is_valid_cotton_key(key, pattern):
                continue
            target = items.setdefault(
                key,
                {
                    "fabric_number": key,
                    "pattern_no": pattern,
                    "fabric_raw": None,
                    "construction": None,
                    "available_meters": None,
                    "composition": None,
                    "unit_price": None,
                },
            )
        if inv.get("unit_price"):
            target["unit_price"] = inv["unit_price"]
        if inv.get("composition"):
            target["composition"] = inv["composition"]
        if pattern and not target.get("pattern_no"):
            target["pattern_no"] = pattern


def is_valid_cotton_key(key: str | None, pattern: str | None = None) -> bool:
    """Reject packing-list / invoice footer rows that are not fabric codes."""
    text = (key or pattern or "").strip()
    if not text:
        return False
    upper = text.upper()
    if any(
        token in upper
        for token in ("METERS", "BALES", "VOLUME", "WEIGHT", "THE END", "CBM", "TOTAL")
    ):
        return False
    if any(char in text for char in ("品", "规", "格")):
        return False
    if re.match(r"^C?\d{5,}(?:-\d+)?(?:BP|-[A-Z]{1,4})?$", text, re.IGNORECASE):
        return True
    if pattern and re.match(r"^C\d{5,}(?:-\d+)?$", str(pattern).strip(), re.IGNORECASE):
        return True
    return False


def parse_canclini_cotton(files: list[Path]) -> list[dict]:
    items: dict[str, dict] = {}
    for path in files:
        if not path.exists():
            print(f"Skipping missing file: {path}", file=sys.stderr)
            continue
        pl = parse_packing_list(path)
        for key, rec in pl.items():
            merged = items.setdefault(key, rec.copy())
            for field in ("pattern_no", "fabric_raw", "construction"):
                if rec.get(field):
                    merged[field] = rec[field]
            if rec.get("available_meters"):
                merged["available_meters"] = (merged.get("available_meters") or 0) + rec["available_meters"]
        apply_invoice_prices(items, parse_invoice(path))

    fabrics: list[dict] = []
    for rec in items.values():
        if not is_valid_cotton_key(rec.get("fabric_number"), rec.get("pattern_no")):
            continue
        construction = rec.get("construction")
        fabrics.append(
            {
                "fabric_number": rec["fabric_number"],
                "composition": rec.get("composition") or "Cotton shirting",
                "color": None,
                "description": f"Canclini cotton shirting — pattern {rec.get('pattern_no') or rec['fabric_number']}",
                "weight_gsm": None,
                "width_cm": None,
                "collection": "HAGAN Cotton Stock (Canclini)",
                "category": "cotton-shirting",
                "unit_price": rec.get("unit_price"),
                "unit": "meters",
                "currency": "USD",
                "is_active": True,
                "stock_status": "in_stock",
                "available_meters": rec.get("available_meters"),
                "source": "canclini-cotton",
                "pattern_no": rec.get("pattern_no"),
                "fabric_raw": rec.get("fabric_raw"),
                "construction": construction,
            }
        )

    fabrics.sort(key=lambda item: item["fabric_number"].lower())
    return fabrics


def merge_fabrics(linen: list[dict], cotton: list[dict]) -> list[dict]:
    by_number: dict[str, dict] = {}
    for fabric in linen:
        by_number[fabric["fabric_number"].lower()] = fabric
    for fabric in cotton:
        by_number[fabric["fabric_number"].lower()] = fabric
    return sorted(by_number.values(), key=lambda item: item["fabric_number"].lower())


def main() -> None:
    linen_path = Path(sys.argv[1]) if len(sys.argv) > 1 else LINEN_XLSX
    cotton_paths = COTTON_FILES

    linen = parse_linen(linen_path) if linen_path.exists() else []
    if not linen and CANCLINI_JSON.exists():
        existing = json.loads(CANCLINI_JSON.read_text(encoding="utf-8"))
        linen = [f for f in existing.get("fabrics", []) if f.get("source", "linen") != "canclini-cotton"]

    cotton = parse_canclini_cotton(cotton_paths)
    fabrics = merge_fabrics(linen, cotton)

    payload = {
        "document_type": "warehouse_stock",
        "supplier": {
            "code": "CANCLINI",
            "name": "Canclini",
            "country": "Italy",
            "is_fabric_supplier": True,
            "lead_time_days": 0,
            "currency": "EUR",
        },
        "price_list_name": "HAGAN Warehouse Stock (Canclini linen + cotton)",
        "imported_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source_file": linen_path.name,
        "source_files": [linen_path.name] + [p.name for p in cotton_paths if p.exists()],
        "fabric_count": len(fabrics),
        "fabrics": fabrics,
    }

    CANCLINI_JSON.parent.mkdir(parents=True, exist_ok=True)
    CANCLINI_JSON.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    priced = sum(1 for f in fabrics if f.get("unit_price"))
    print(
        f"Canclini warehouse stock: {len(linen)} linen + {len(cotton)} cotton = {len(fabrics)} fabrics "
        f"({priced} with list price, admin-only) -> {CANCLINI_JSON}"
    )


if __name__ == "__main__":
    main()
