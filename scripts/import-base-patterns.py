#!/usr/bin/env python3
"""
Import base-pattern measurement Excel files into src/data/pattern-library.json.

Handles the two spreadsheet formats the pattern team uses:
  A) Standard FR measurement spec — "Measurement point" header row, sizes across
     the top, optional Remarks column, "Special instruction:" footer.
  B) Hagan tech-pack (e.g. Massimo Linen Short 0197) — points with size ref
     (diagram code), tolerance, grading increment; unit cm.

Empty size columns/rows are preserved as nulls (sparse grids are real data).
Re-running upserts by deterministic base-pattern id, so tomorrow's GL files can
be imported with:  python3 scripts/import-base-patterns.py --folder <dir> \
    --brand-id gliani --brand-code GL

Usage:
  python3 scripts/import-base-patterns.py --folder "/path/to/Base Patterns " \
      [--brand-id fouad-rahme] [--brand-code FR] [--dry-run]

After importing, push the document to Supabase with:
  npm run db:migrate-json   (or let the app save it on first library write)
"""

import argparse
import datetime
import json
import os
import re
import sys
import unicodedata

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip3 install openpyxl")

LIBRARY_PATH = os.path.join(os.path.dirname(__file__), "..", "src", "data", "pattern-library.json")

# ---------------------------------------------------------------------------
# Hygiene + canonical naming
# ---------------------------------------------------------------------------

# Misspellings / brand hygiene fixes seen in the team's files.
CUT_FAMILY_FIXES = {
    "boggy": "Boggi",
    "boggi": "Boggi",
    "suit suplly": "Suit Supply",
    "suit supply": "Suit Supply",
    "comfort": "Comfort",
    "massimo dutty": "Massimo",
    "massimo dutti": "Massimo",
    "massimo": "Massimo",
    "hugo boss": "Hugo Boss",
    "brand boss": "Hugo Boss",
    "boss": "Hugo Boss",
    "camicissima": "Camicissima",
    "luca faloni": "Luca Faloni",
}

# Alias -> canonical measurement point name (per the team's own equivalences).
POINT_ALIASES = {
    "1/2 bottom width": "1/2 Hem Width",
    "total length (nps)": "Total Length (HNP)",
    "cuff width": "Cuff Height",
}

# Trim points that are constant across sizes even when the grid is sparse.
KNOWN_TRIM_POINTS = {
    "collar height",
    "collar point",
    "band height",
    "cuff height",
    "cuff width",
    "slv placket length",
    "waist band height",
    "bottom hem stitch height",
    "fly width",
}

GARMENT_HINTS = [
    ("overshirt", "overshirt"),
    ("t-shirt", "t-shirt"),
    ("tee-shirt", "t-shirt"),
    ("tee shirt", "t-shirt"),
    ("short pant", "shorts"),
    ("linen short", "shorts"),
    ("shorts", "shorts"),
    ("chino", "trouser"),
    ("trouser", "trouser"),
    ("jacket", "jacket"),
    ("kurta", "shirt"),
    ("shirt", "shirt"),
    ("pant", "trouser"),
    ("thobe", "thobe"),
    # Bare "short" last — Suit Supply Short trousers also contain the word.
    ("short", "shorts"),
]

# House blank measurement templates — filename → library garment_type key(s).
TEMPLATE_FILE_GARMENTS = {
    "jacket.xlsx": ["jacket"],
    "shirt-over shirt.xlsx": ["shirt", "overshirt"],
    "short.xlsx": ["shorts"],
    "t-shirt measurement.xlsx": ["t-shirt"],
    "thobe.xlsx": ["thobe"],
    "trouser.xlsx": ["trouser"],
}

TEMPLATE_CUT_FAMILY = "House Template"

# Trial/sewing column headers on blank templates — not graded size runs.
_TRIAL_HEADER = re.compile(
    r"^(size|sewing|trial|adjustment|final|pattern\s*size|client\s*body|"
    r"size\{cm\}|sewimg|delevery|order\s*date|fabric\s*color)$",
    re.I,
)


def is_trial_workflow_grid(sizes):
    """True when every header column is a trial/sewing label, not a size run."""
    if not sizes:
        return False
    return all(_TRIAL_HEADER.match(clean(size)) for size in sizes)


def template_garments_for_file(filename, description=""):
    """Map a blank-template workbook to one or more library garment keys."""
    key = filename.lower()
    if key in TEMPLATE_FILE_GARMENTS:
        return list(TEMPLATE_FILE_GARMENTS[key])
    text = f"{filename} {description}".lower()
    found = []
    for hint, garment in GARMENT_HINTS:
        if hint in text and garment not in found:
            found.append(garment)
    return found or ["unknown"]


def extract_description_from_rows(rows, header_idx):
    """Read the Description metadata cell (e.g. 'Jacket', 'Shirt / Overshirt')."""
    for row in rows[:header_idx]:
        cells = [clean(cell) for cell in (row or [])]
        for j, cell in enumerate(cells):
            if cell.lower().rstrip(":") == "description":
                return clean_name(next((c for c in cells[j + 1 :] if c and c != ":"), None))
    return None


def now_iso():
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.000Z")


def slugify(text):
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", text.lower())).strip("-")


def clean(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def clean_name(value):
    """Strip leading label junk: ': Brand Boss' → 'Brand Boss'."""
    text = clean(value)
    text = re.sub(r"^[:\-\s]+", "", text)
    return clean(text)


def canonical_point_name(raw_name):
    return POINT_ALIASES.get(clean(raw_name).lower(), clean(raw_name))


def to_number(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        # Mixed numbers: "4 9/16"
        mixed = re.fullmatch(r"(\d+)\s+(\d+)\s*/\s*(\d+)", text)
        if mixed:
            whole, num, den = (int(mixed.group(i)) for i in (1, 2, 3))
            return whole + (num / den) if den else float(whole)
        frac = re.fullmatch(r"(\d+)\s*/\s*(\d+)", text)
        if frac:
            num, den = int(frac.group(1)), int(frac.group(2))
            return num / den if den else None
        try:
            return float(text)
        except ValueError:
            return None
    return None


def normalize_size(value):
    """'S- 22' -> 'S-22'; 48.0 -> '48'; 'L- 90' -> 'L-90'."""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return str(int(value)) if float(value).is_integer() else str(value)
    text = clean(value)
    text = re.sub(r"([A-Za-z])\s*-\s*", r"\1-", text)
    return text


def detect_garment(*texts):
    for text in texts:
        lowered = clean(text).lower()
        for hint, garment in GARMENT_HINTS:
            if hint in lowered:
                return garment
    return None


_VARIANT_NOISE = re.compile(
    r"\b(shirt|trouser|trousers|pant|pants|jacket|shorts|chino|kurta|"
    r"linen|cotton|wool|mens|men'?s|for|all|measurement|spec)\b",
    re.I,
)


def tidy_variant(raw):
    """Keep fit/length variants (Fitted, Slim, Long); drop fabric/garment noise."""
    text = _VARIANT_NOISE.sub("", clean(raw) or "")
    text = clean(re.sub(r"\s+", " ", text))
    if not text or text.lower() in {"fit", "fits"}:
        return None
    # "Slim Fit" → Slim
    text = re.sub(r"\bfit\b", "", text, flags=re.I)
    text = clean(text)
    return text.title() if text else None


def parse_family_and_variant(raw_name):
    """' Suit Suplly (Short)' -> ('Suit Supply', 'Short'); 'BOGGY' -> ('Boggi', None)."""
    name = clean_name(raw_name)
    variant = None
    match = re.search(r"\(([^)]+)\)\s*$", name)
    if match:
        variant = tidy_variant(match.group(1))
        name = clean(name[: match.start()])
    lowered = name.lower()
    for key in sorted(CUT_FAMILY_FIXES, key=len, reverse=True):
        if lowered == key or lowered.startswith(key + " ") or lowered.startswith(key + "("):
            family = CUT_FAMILY_FIXES[key]
            rest = clean(name[len(key) :])
            if rest and not variant:
                variant = tidy_variant(rest)
            return family, variant
    return name.title() if name else None, variant


def infer_unit(sizes, points, default="in"):
    """Collar charts with body lengths ≥55 are metric; otherwise keep default."""
    numeric_sizes = []
    for size in sizes:
        try:
            numeric_sizes.append(float(size))
        except ValueError:
            return default
    if not numeric_sizes or min(numeric_sizes) < 34 or max(numeric_sizes) > 52:
        return default
    lengthish = [
        v
        for point in points
        for v in point["values"].values()
        if v is not None and "length" in point["name"].lower()
    ]
    if lengthish and max(lengthish) >= 55:
        return "cm"
    return default


def detect_fabric(*texts):
    for text in texts:
        lowered = clean(text).lower()
        if "linen" in lowered:
            return "linen"
        if "cotton" in lowered:
            return "cotton"
        if "wool" in lowered:
            return "wool"
    return None


# ---------------------------------------------------------------------------
# Format A: standard FR measurement spec
# ---------------------------------------------------------------------------

def parse_standard_sheet(ws, filename):
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]

    header_idx = None
    for i, row in enumerate(rows):
        if clean(row[0] if row else "").lower().startswith("measurement point"):
            header_idx = i
            break
    if header_idx is None:
        return None

    header = rows[header_idx]
    remarks_col = None
    for j, cell in enumerate(header):
        if clean(cell).lower() == "remarks":
            remarks_col = j
            break

    size_cols = []  # (col_index, size)
    limit = remarks_col if remarks_col is not None else len(header)
    for j in range(1, limit):
        size = normalize_size(header[j])
        if size:
            size_cols.append((j, size))
    sizes = [size for _, size in size_cols]

    # Metadata rows above the header.
    pattern_name = None
    description = None
    style_code = None
    extra_notes = []
    title = clean(rows[0][0] if rows and rows[0] else "")
    for row in rows[:header_idx]:
        cells = [clean(cell) for cell in (row or [])]
        for j, cell in enumerate(cells):
            lowered = cell.lower().rstrip(":")
            if lowered == "name":
                pattern_name = clean_name(
                    next((c for c in cells[j + 1 :] if c and c != ":"), None)
                )
            if lowered == "description":
                description = clean_name(
                    next((c for c in cells[j + 1 :] if c and c != ":"), None)
                )
            if lowered in ("code", "model", "pattern  ref", "pattern ref"):
                style_code = clean_name(
                    next((c for c in cells[j + 1 :] if c and c != ":"), None)
                ) or style_code
            if "fabric" in lowered and ("code" in lowered or "name" in lowered):
                code = clean_name(next((c for c in cells[j + 1 :] if c and c != ":"), None))
                if code:
                    extra_notes.append(f"Fabric code: {code}")

    points = []
    special_instructions = None
    for row in rows[header_idx + 1 :]:
        cells = row or []
        first = clean(cells[0] if cells else "")
        if first.lower().startswith("special instruction"):
            trailing = clean(first.split(":", 1)[1]) if ":" in first else ""
            rest = " ".join(clean(c) for c in cells[1:] if clean(c))
            special_instructions = clean(f"{trailing} {rest}") or None
            break
        if first.lower().startswith(("total ordered", "remarks")):
            continue
        if not first:
            continue

        values = {}
        for j, size in size_cols:
            values[size] = to_number(cells[j]) if j < len(cells) else None
        remark = clean(cells[remarks_col]) if remarks_col is not None and remarks_col < len(cells) else ""
        non_null = [v for v in values.values() if v is not None]
        fully_constant = len(non_null) == len(sizes) and len(set(non_null)) == 1 and len(sizes) > 1
        canonical = canonical_point_name(first)
        points.append(
            {
                "point_id": slugify(canonical),
                "name": clean(first),
                "remark": remark or None,
                "is_graded": not (fully_constant or canonical.lower() in KNOWN_TRIM_POINTS),
                "tolerance": None,
                "grading_increment": None,
                "diagram_code": None,
                "values": values,
            }
        )

    family, variant = parse_family_and_variant(pattern_name or "")
    if not variant and description:
        variant = tidy_variant(description)
    garment = (
        detect_garment(description or "", title, pattern_name or "", filename) or "unknown"
    )
    fabric = detect_fabric(
        pattern_name or "",
        description or "",
        style_code or "",
        filename,
        *extra_notes,
    )
    unit = infer_unit(sizes, points, default="in")

    return {
        "cut_family": family or "Unknown",
        "cut_variant": variant,
        "garment_type": garment,
        "unit": unit,
        "sizes": sizes,
        "points": points,
        "style_code": clean(style_code) or None,
        "fabric": fabric,
        "season": None,
        "special_instructions": special_instructions,
        "notes": "; ".join(extra_notes) or None,
        "display_name": clean(pattern_name) or None,
    }


# ---------------------------------------------------------------------------
# Format C: Hagan tee-shirt grid (MEASUREMENT POINT + alpha sizes, often blank)
# ---------------------------------------------------------------------------

def parse_hagan_tee_sheet(ws, filename):
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]

    header_idx = None
    point_col = None
    ref_col = None
    grading_col = None
    tol_col = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row or []):
            lowered = clean(cell).lower()
            if lowered.startswith("measurement point"):
                header_idx = i
                point_col = j
            if lowered in ("size ref", "ref") or (len(lowered) == 1 and lowered.isalpha()):
                ref_col = j if ref_col is None else ref_col
            if lowered.startswith("development grade"):
                grading_col = j
            if lowered.startswith("tol"):
                tol_col = j
        if header_idx is not None:
            break
    if header_idx is None or point_col is None:
        return None

    header = rows[header_idx]
    ref_col = point_col + 1
    if ref_col < len(header) and clean(header[ref_col]):
        lowered_ref = clean(header[ref_col]).lower()
        if lowered_ref.startswith("development") or lowered_ref in ("s", "m", "l", "xl", "xxl"):
            ref_col = None

    size_cols = []
    start = ref_col + 1 if ref_col is not None else point_col + 1
    for j in range(start, len(header)):
        label = clean(header[j])
        lowered = label.lower()
        if lowered.startswith("development grade") or lowered.startswith("tol"):
            break
        size = normalize_size(header[j])
        if size and not lowered.startswith("measurement"):
            size_cols.append((j, size))

    points = []
    for row in rows[header_idx + 1 :]:
        cells = row or []
        name = clean(cells[point_col] if point_col < len(cells) else "")
        if not name:
            continue
        diagram = clean(cells[ref_col]) if ref_col is not None and ref_col < len(cells) else None
        grading = (
            to_number(cells[grading_col])
            if grading_col is not None and grading_col < len(cells)
            else None
        )
        tolerance = (
            to_number(cells[tol_col]) if tol_col is not None and tol_col < len(cells) else None
        )
        values = {}
        for j, size in size_cols:
            values[size] = to_number(cells[j]) if j < len(cells) else None
        canonical = canonical_point_name(name)
        points.append(
            {
                "point_id": slugify(canonical),
                "name": name,
                "remark": None,
                "is_graded": bool(grading) and grading != 0,
                "tolerance": tolerance,
                "grading_increment": grading,
                "diagram_code": diagram or None,
                "values": values,
            }
        )

    description = None
    for row in rows[: header_idx + 1]:
        for cell in row or []:
            text = clean(cell)
            if "tee-shirt" in text.lower() or "t-shirt" in text.lower():
                description = text
                break

    return {
        "cut_family": TEMPLATE_CUT_FAMILY,
        "cut_variant": None,
        "garment_type": "t-shirt",
        "unit": "cm",
        "sizes": [size for _, size in size_cols],
        "points": points,
        "style_code": None,
        "fabric": detect_fabric(description or "", filename),
        "season": None,
        "special_instructions": None,
        "notes": description,
        "display_name": "T-shirt measurement template",
    }


def normalize_blank_template(parsed, filename):
    """Convert a parsed trial-column grid into a house blank template base."""
    description = parsed.get("_description") or parsed.get("notes") or ""
    garments = template_garments_for_file(filename, description)
    unit = parsed["unit"]
    if any(g in ("shorts", "trouser", "thobe", "t-shirt") for g in garments):
        if parsed["unit"] == "in" and any(
            "cm" in clean(s).lower() for s in (parsed.get("sizes") or [])
        ):
            unit = "cm"
        elif parsed["garment_type"] in ("shorts", "trouser", "thobe", "t-shirt"):
            unit = parsed.get("_template_unit") or parsed["unit"]

    blank_points = []
    for point in parsed["points"]:
        blank_points.append(
            {
                **point,
                "values": {},
                "is_graded": point.get("is_graded", True),
            }
        )

    results = []
    for garment in garments:
        results.append(
            {
                **parsed,
                "cut_family": TEMPLATE_CUT_FAMILY,
                "cut_variant": None,
                "garment_type": garment,
                "unit": unit,
                "sizes": [],
                "points": [dict(p) for p in blank_points],
                "display_name": f"House {garment.replace('-', ' ').title()} template",
            }
        )
    return results


# ---------------------------------------------------------------------------
# Format B: Hagan tech-pack (size ref / tolerance / grading, cm)
# ---------------------------------------------------------------------------

def parse_techpack_sheet(ws, filename):
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]

    header_idx = None
    ref_col = None
    for i, row in enumerate(rows):
        for j, cell in enumerate(row or []):
            if clean(cell).lower().startswith("size ref"):
                header_idx, ref_col = i, j
                break
        if header_idx is not None:
            break
    if header_idx is None:
        return None

    header = rows[header_idx]
    point_col = ref_col - 1
    tol_col = None
    grading_col = None
    for j, cell in enumerate(header):
        lowered = clean(cell).lower()
        if lowered.startswith("tole"):
            tol_col = j
        if lowered == "grading":
            grading_col = j

    size_cols = []
    start = (tol_col if tol_col is not None else ref_col) + 1
    stop = grading_col if grading_col is not None else len(header)
    for j in range(start, stop):
        size = normalize_size(header[j])
        if size:
            size_cols.append((j, size))
    sizes = [size for _, size in size_cols]

    # Metadata: customer / style / model / season live above the header row.
    customer = None
    style_name = None
    style_code = None
    season = None
    for row in rows[: header_idx + 1]:
        cells = [clean(cell) for cell in (row or [])]
        for j, cell in enumerate(cells):
            lowered = cell.lower()
            if lowered.startswith("customer") and j + 1 < len(cells):
                customer = next((c for c in cells[j + 1 :] if c), None)
            if lowered == "style name" and j + 1 < len(cells):
                style_name = next((c for c in cells[j + 1 :] if c), None)
            if lowered.startswith("model") and j + 1 < len(cells):
                style_code = next((c for c in cells[j + 1 :] if c), None)
            if re.fullmatch(r"(SS|AW|FW)\d{2}", cell):
                season = cell

    points = []
    for row in rows[header_idx + 1 :]:
        cells = row or []
        name = clean(cells[point_col] if point_col < len(cells) else "")
        if not name:
            continue
        values = {}
        for j, size in size_cols:
            values[size] = to_number(cells[j]) if j < len(cells) else None
        grading = to_number(cells[grading_col]) if grading_col is not None and grading_col < len(cells) else None
        canonical = canonical_point_name(name)
        points.append(
            {
                "point_id": slugify(canonical),
                "name": name,
                "remark": None,
                "is_graded": bool(grading) and grading != 0,
                "tolerance": to_number(cells[tol_col]) if tol_col is not None and tol_col < len(cells) else None,
                "grading_increment": grading,
                "diagram_code": clean(cells[ref_col]) if ref_col < len(cells) else None or None,
                "values": values,
            }
        )
    for point in points:
        if not point["diagram_code"]:
            point["diagram_code"] = None

    family, variant = parse_family_and_variant(customer or "")
    if not variant and style_name and "chino" in clean(style_name).lower():
        variant = "Chino"
    garment = detect_garment(style_name or "", filename) or "unknown"
    fabric_bits = []
    for row in rows[: header_idx + 1]:
        for cell in row or []:
            text = clean(cell)
            if "cotton" in text.lower() or "linen" in text.lower() or "wool" in text.lower():
                fabric_bits.append(text)
    fabric = detect_fabric(style_name or "", filename, *fabric_bits)

    return {
        "cut_family": family or "Unknown",
        "cut_variant": variant,
        "garment_type": garment,
        "unit": "cm",
        "sizes": sizes,
        "points": points,
        "style_code": clean(style_code) or None,
        "fabric": fabric,
        "season": season,
        "special_instructions": None,
        "notes": clean(style_name) or None,
        "display_name": clean(style_name) or None,
    }


# ---------------------------------------------------------------------------
# Assembly
# ---------------------------------------------------------------------------

def parse_workbook(path, templates=False):
    wb = openpyxl.load_workbook(path, data_only=True)
    filename = os.path.basename(path)
    for ws in wb.worksheets:
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        flat = [clean(cell.value).lower() for row in ws.iter_rows() for cell in row]

        hagan_tee = filename.lower().startswith("t-shirt") or filename.lower().startswith("tee")
        if not hagan_tee:
            for row in rows:
                for j, cell in enumerate(row or []):
                    if clean(cell).lower().startswith("measurement point") and j > 0:
                        headers = {clean(c).upper() for c in (row or []) if clean(c)}
                        if {"S", "M", "L"}.issubset(headers):
                            hagan_tee = True
                        break
                if hagan_tee:
                    break

        if hagan_tee:
            parsed = parse_hagan_tee_sheet(ws, filename)
        elif any(text.startswith("size ref") for text in flat):
            parsed = parse_techpack_sheet(ws, filename)
        else:
            parsed = parse_standard_sheet(ws, filename)
        if not parsed or not parsed["points"]:
            continue
        if templates and (
            is_trial_workflow_grid(parsed["sizes"])
            or filename.lower() in TEMPLATE_FILE_GARMENTS
            or parsed.get("garment_type") == "t-shirt"
        ):
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            header_idx = next(
                (
                    i
                    for i, row in enumerate(rows)
                    if clean(row[0] if row else "").lower().startswith("measurement point")
                ),
                0,
            )
            parsed["_description"] = extract_description_from_rows(rows, header_idx)
            if "cm" in filename.lower() or any(
                "cm" in clean(s).lower() for s in parsed.get("sizes", [])
            ):
                parsed["_template_unit"] = "cm"
            return normalize_blank_template(parsed, filename)
        if parsed and parsed["points"]:
            return [parsed]
    return []


def base_pattern_id(brand_code, parsed):
    parts = [
        "bp",
        brand_code.lower(),
        slugify(parsed["cut_family"]),
        slugify(parsed["garment_type"]),
    ]
    if parsed["cut_variant"]:
        parts.append(slugify(parsed["cut_variant"]))
    # Disambiguate multiple bases of the same family+garment (Massimo cotton vs linen).
    if parsed["fabric"]:
        parts.append(slugify(parsed["fabric"]))
    if parsed["style_code"]:
        parts.append(slugify(parsed["style_code"]))
    return "-".join(parts)


def display_name(brand_code, parsed):
    bits = [parsed["cut_family"], parsed["garment_type"].capitalize()]
    if parsed["fabric"]:
        bits.insert(1, parsed["fabric"].capitalize())
    if parsed["cut_variant"]:
        bits.append(f"({parsed['cut_variant']})")
    if parsed["style_code"]:
        bits.append(parsed["style_code"])
    return " ".join(bits)


def merge_dictionary(dictionary, base):
    by_id = {entry["id"]: entry for entry in dictionary}
    for point in base["points"]:
        canonical = canonical_point_name(point["name"])
        entry = by_id.get(point["point_id"])
        if entry is None:
            entry = {
                "id": point["point_id"],
                "name": canonical,
                "aliases": [],
                "garment_types": [],
            }
            dictionary.append(entry)
            by_id[point["point_id"]] = entry
        display = clean(point["name"])
        if display != entry["name"] and display not in entry["aliases"]:
            entry["aliases"].append(display)
        if base["garment_type"] not in entry["garment_types"]:
            entry["garment_types"].append(base["garment_type"])


def detect_brand_from_path(path):
    """Infer FR/GL from folder segments or filename prefixes. Returns (brand_id, brand_code) or None."""
    parts = [p.lower() for p in path.replace("\\", "/").split("/")]
    joined = "/".join(parts)
    filename = parts[-1] if parts else ""
    # Prefer explicit house-brand folder markers over filename codes.
    if any(
        token in parts
        for token in ("gliani", "gliani beirut", "gl")
    ) or "/gliani" in joined or "/gl/" in f"/{joined}/" or "gliani" in filename:
        return ("gliani", "GL")
    if any(
        token in parts
        for token in ("foad rahme", "fouad rahme", "foad rahme spec", "fr", "fr beirut")
    ) or "/foad" in joined or "/fouad" in joined or "/fr/" in f"/{joined}/":
        return ("fouad-rahme", "FR")
    if re.search(r"(^|[\s_-])gl([\s_-]|$)", filename) or "-gl-" in filename:
        return ("gliani", "GL")
    if re.search(r"(^|[\s_-])fr([\s_-]|$)", filename) or "-fr-" in filename:
        return ("fouad-rahme", "FR")
    return None


def iter_xlsx_files(folder, recursive):
    if not recursive:
        for filename in sorted(os.listdir(folder)):
            if filename.lower().endswith(".xlsx") and not filename.startswith("~$"):
                yield os.path.join(folder, filename)
        return
    for dirpath, dirnames, filenames in os.walk(folder, followlinks=True):
        # Skip junk / lock / copy dumps (prune so we don't descend further)
        dirnames[:] = [
            d
            for d in dirnames
            if d.lower() not in {"copy", "marker", "old", "new folder", ".ds_store"}
            and not d.startswith(".")
        ]
        low_parts = {p.lower() for p in dirpath.replace("\\", "/").split("/")}
        if low_parts & {"copy", "marker", "old", "new folder"}:
            continue
        for filename in sorted(filenames):
            if not filename.lower().endswith(".xlsx") or filename.startswith("~$"):
                continue
            if "copy" in filename.lower():
                continue
            yield os.path.join(dirpath, filename)


def filled_cell_count(parsed):
    return sum(1 for p in parsed["points"] for v in p["values"].values() if v is not None)


def cell_count(parsed):
    return sum(len(p["values"]) for p in parsed["points"])


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--folder", required=True, help="Folder containing the .xlsx files")
    parser.add_argument("--brand-id", default="fouad-rahme", help="Factory brand id (fouad-rahme, gliani)")
    parser.add_argument("--brand-code", default="FR", help="House brand code (FR, GL)")
    parser.add_argument("--recursive", action="store_true", help="Walk subfolders for .xlsx files")
    parser.add_argument(
        "--detect-brand",
        action="store_true",
        help="Infer FR/GL from path (falls back to --brand-id/--brand-code)",
    )
    parser.add_argument("--min-sizes", type=int, default=3, help="Skip grids with fewer size columns")
    parser.add_argument(
        "--min-fill-ratio",
        type=float,
        default=0.35,
        help="Skip sparse/per-size sheets below this filled/cells ratio",
    )
    parser.add_argument(
        "--templates",
        action="store_true",
        help="Import house blank measurement templates (trial-column grids, no size values)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse and report without writing")
    args = parser.parse_args()

    folder = args.folder
    if not os.path.isdir(folder):
        sys.exit(f"Folder not found: {folder!r}")

    library_path = os.path.abspath(LIBRARY_PATH)
    with open(library_path) as fh:
        library = json.load(fh)
    library.setdefault("dictionary", [])
    library.setdefault("base_patterns", [])
    library.setdefault("client_patterns", [])
    # Never drop remote/local client patterns during base import.
    preserved_client_patterns = library["client_patterns"]

    existing_by_id = {base["id"]: base for base in library["base_patterns"]}
    timestamp = now_iso()
    imported = []
    skipped = []
    candidates_by_id = {}  # base_id -> (filled, base_dict, rel_source)

    for path in iter_xlsx_files(folder, args.recursive):
        filename = os.path.basename(path)
        rel = os.path.relpath(path, folder)
        try:
            parsed_list = parse_workbook(path, templates=args.templates)
        except Exception as exc:
            skipped.append((rel, f"parse error: {exc}"))
            continue
        if not parsed_list:
            skipped.append((rel, "unrecognized layout"))
            continue

        brand_id, brand_code = args.brand_id, args.brand_code.upper()
        if args.detect_brand:
            detected = detect_brand_from_path(path)
            if detected:
                brand_id, brand_code = detected

        for parsed in parsed_list:
            if not parsed or not parsed["points"]:
                continue

            n_sizes = len(parsed["sizes"])
            filled = filled_cell_count(parsed)
            cells = cell_count(parsed) or 1
            fill_ratio = filled / cells
            is_template = args.templates and parsed["cut_family"] == TEMPLATE_CUT_FAMILY
            if not is_template:
                if n_sizes < args.min_sizes:
                    skipped.append((rel, f"only {n_sizes} size column(s)"))
                    continue
                if fill_ratio < args.min_fill_ratio:
                    skipped.append((rel, f"sparse fill {filled}/{cells} ({fill_ratio:.0%})"))
                    continue

            base_id = base_pattern_id(brand_code, parsed)
            previous = existing_by_id.get(base_id)
            # Prefer denser grids when multiple files collapse to the same id.
            prior_candidate = candidates_by_id.get(base_id)
            if prior_candidate and prior_candidate[0] >= filled and not is_template:
                skipped.append((rel, f"duplicate of denser {prior_candidate[2]}"))
                continue

            display = parsed.get("display_name") or display_name(brand_code, parsed)
            base = {
                "id": base_id,
                "house_brand_id": brand_id,
                "house_brand_code": brand_code,
                "cut_family": parsed["cut_family"],
                "garment_type": parsed["garment_type"],
                "cut_variant": parsed["cut_variant"],
                "name": display,
                "unit": parsed["unit"],
                "sizes": parsed["sizes"],
                "points": parsed["points"],
                "style_code": parsed["style_code"],
                "fabric": parsed["fabric"],
                "season": parsed["season"],
                "special_instructions": parsed["special_instructions"],
                "physical_pattern_kept": previous.get("physical_pattern_kept", False) if previous else False,
                "physical_pattern_location": previous.get("physical_pattern_location") if previous else None,
                "files": previous.get("files", []) if previous else [],
                "source_file": rel if args.recursive else filename,
                "notes": parsed["notes"],
                "created_at": previous.get("created_at", timestamp) if previous else timestamp,
                "updated_at": timestamp,
            }
            score = len(base["points"]) if is_template else filled
            candidates_by_id[base_id] = (score, base, rel)
            size_label = f"{n_sizes} sizes" if n_sizes else "template (no sizes)"
            fill_label = f"{len(base['points'])} points" if is_template else f"{filled}/{cells} cells filled"
            print(
                f"  OK [{brand_code}] {base_id}\n"
                f"     {base['name']} | {size_label} | {fill_label} | src={rel}"
            )

    for base_id, (filled, base, _rel) in sorted(candidates_by_id.items()):
        existing_by_id[base_id] = base
        merge_dictionary(library["dictionary"], base)
        imported.append(base)

    library["base_patterns"] = sorted(existing_by_id.values(), key=lambda b: b["id"])
    library["client_patterns"] = preserved_client_patterns
    library["updated_at"] = timestamp

    print(
        f"\nImported/updated {len(imported)} base pattern(s); library now has "
        f"{len(library['base_patterns'])} base(s), {len(library['dictionary'])} dictionary point(s), "
        f"{len(library['client_patterns'])} client pattern(s)."
    )
    print(f"Skipped {len(skipped)} file(s).")
    for rel, reason in skipped[:40]:
        print(f"  SKIP {rel}: {reason}")
    if len(skipped) > 40:
        print(f"  ... and {len(skipped) - 40} more")

    if args.dry_run:
        print("Dry run — nothing written.")
        return

    with open(library_path, "w") as fh:
        json.dump(library, fh, indent=2, ensure_ascii=False)
        fh.write("\n")
    print(f"Wrote {library_path}")


if __name__ == "__main__":
    main()
