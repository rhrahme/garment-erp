#!/usr/bin/env python3
"""Import HAGAN wool warehouse stock from Wool Stock .xlsx."""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Desktop" / "Fabrics" / "Wool Stock" / "Wool Stock .xlsx"
OUT_PATH = ROOT / "src" / "data" / "suppliers" / "wool-stock.json"

MERGE_FIELDS = ("unit_price", "available_meters", "width_cm", "color", "composition", "weight_gsm")


def normalize_fabric_number(value) -> str | None:
    """Strip vendor prefixes (TT, KKTT) to match short warehouse codes."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"reference", "total"}:
        return None
    normalized = re.sub(r"^(?:KKTT|TT)+", "", text, flags=re.IGNORECASE)
    return normalized or text


def parse_weight(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[\d.]+", str(value))
    return float(match.group(0)) if match else None


def is_valid_reference(value) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text) and text.lower() not in {"reference", "total"}


def build_quality_lookup(rows: list[tuple]) -> dict[str, dict]:
    """Collect composition/weight by quality number from rows that have specs."""
    lookup: dict[str, dict] = {}
    for reference, quality_no, composition, weight in rows:
        if not quality_no:
            continue
        quality = str(quality_no).strip()
        comp = str(composition).strip() if composition else None
        weight_gsm = parse_weight(weight)
        if not comp and weight_gsm is None:
            continue
        entry = lookup.setdefault(quality, {})
        if comp:
            entry["composition"] = comp
        if weight_gsm is not None:
            entry["weight_gsm"] = weight_gsm
    return lookup


def parse_sheet(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["All References"]
    raw_rows: list[tuple] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        reference, quality_no, composition, weight = (row + (None, None, None, None))[:4]
        if not is_valid_reference(reference):
            continue
        raw_rows.append((reference, quality_no, composition, weight))

    wb.close()
    quality_lookup = build_quality_lookup(raw_rows)

    fabrics: list[dict] = []
    seen: set[str] = set()

    for reference, quality_no, composition, weight in raw_rows:
        raw_reference = str(reference).strip()
        fabric_number = normalize_fabric_number(raw_reference)
        if not fabric_number:
            continue

        key = fabric_number.lower()
        if key in seen:
            continue
        seen.add(key)

        quality = str(quality_no).strip() if quality_no else None
        quality_meta = quality_lookup.get(quality or "", {})

        comp = str(composition).strip() if composition else quality_meta.get("composition")
        weight_gsm = parse_weight(weight)
        if weight_gsm is None:
            weight_gsm = quality_meta.get("weight_gsm")

        weight_label = f"{int(weight_gsm)} g/m" if weight_gsm else "spec pending"
        description_parts = ["Wool stock"]
        if quality:
            description_parts.append(f"quality {quality}")
        if comp:
            description_parts.append(comp)
        description_parts.append(f"({weight_label})")

        fabric: dict = {
            "fabric_number": fabric_number,
            "composition": comp,
            "color": None,
            "description": " — ".join(description_parts),
            "weight_gsm": weight_gsm,
            "width_cm": None,
            "collection": "HAGAN Wool Stock",
            "category": quality,
            "unit_price": None,
            "unit": "meters",
            "currency": "EUR",
            "is_active": True,
            "stock_status": "in_stock",
            "available_meters": None,
        }
        if fabric_number.upper() != raw_reference.upper():
            fabric["fabric_raw"] = raw_reference

        fabrics.append(fabric)

    fabrics.sort(key=lambda item: item["fabric_number"].lower())
    return fabrics


def normalize_existing_key(fabric: dict) -> str:
    return (normalize_fabric_number(fabric.get("fabric_number")) or fabric["fabric_number"]).lower()


def merge_with_existing(imported: list[dict], existing_path: Path) -> list[dict]:
    """Merge Excel import with prior catalog; preserve admin fields when Excel lacks them."""
    if not existing_path.exists():
        return imported

    existing_payload = json.loads(existing_path.read_text(encoding="utf-8"))
    by_number: dict[str, dict] = {}

    for fabric in existing_payload.get("fabrics", []):
        key = normalize_existing_key(fabric)
        normalized = normalize_fabric_number(fabric.get("fabric_number")) or fabric["fabric_number"]
        merged = {**fabric, "fabric_number": normalized}
        if normalized != fabric.get("fabric_number"):
            merged["fabric_raw"] = fabric.get("fabric_raw") or fabric["fabric_number"]
        by_number[key] = merged

    for fabric in imported:
        key = fabric["fabric_number"].lower()
        prior = by_number.get(key)
        if not prior:
            by_number[key] = fabric
            continue
        merged = {**prior, **fabric}
        for field in MERGE_FIELDS:
            if fabric.get(field) is None and prior.get(field) is not None:
                merged[field] = prior[field]
        if fabric.get("fabric_raw"):
            merged["fabric_raw"] = fabric["fabric_raw"]
        elif prior.get("fabric_raw"):
            merged["fabric_raw"] = prior["fabric_raw"]
        by_number[key] = merged

    return sorted(by_number.values(), key=lambda item: item["fabric_number"].lower())


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    imported = parse_sheet(xlsx_path)
    fabrics = merge_with_existing(imported, OUT_PATH)
    payload = {
        "document_type": "warehouse_stock",
        "supplier": {
            "code": "WOOL-STOCK",
            "name": "Wool Stock",
            "country": None,
            "is_fabric_supplier": True,
            "lead_time_days": 0,
            "currency": "EUR",
        },
        "price_list_name": "HAGAN Wool Stock",
        "imported_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source_file": xlsx_path.name,
        "fabric_count": len(fabrics),
        "fabrics": fabrics,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    normalized = sum(1 for f in fabrics if f.get("fabric_raw"))
    print(
        f"Wool stock: {len(fabrics)} fabrics ({normalized} with raw vendor prefix, "
        f"prices admin-only) -> {OUT_PATH}"
    )


if __name__ == "__main__":
    main()
