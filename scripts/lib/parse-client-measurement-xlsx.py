#!/usr/bin/env python3
"""Parse FR client measurement Excel sheets into JSON (stdout).

Maps Size / Sample column -> base_value + target_value.
Maps Trial - N columns -> trial_values[N].
Maps Final column -> final_value.
Also extracts pattern_ref, fabric code, description, unit hint, special instructions.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def clean(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = clean(value).replace(",", ".")
    if (not text) or text in {"-", "n/a", "na"} or (len(text) == 1 and ord(text[0]) in (0x2013, 0x2014)):
        return None
    # fractions like 6 5/8
    m = re.fullmatch(r"(\d+)\s+(\d+)/(\d+)", text)
    if m:
        whole, num, den = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return whole + (num / den if den else 0)
    m = re.fullmatch(r"(\d+)/(\d+)", text)
    if m:
        num, den = int(m.group(1)), int(m.group(2))
        return num / den if den else None
    try:
        return float(text)
    except ValueError:
        return None


def slugify(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return s or "point"


def parse_md_short_sheet(path: Path, rows: list) -> dict | None:
    """Factory MD-SHORT style: names in col E, size values under a size label (e.g. 2XL)."""
    header_idx = None
    for i, row in enumerate(rows):
        cells = [clean(c) for c in (row or [])]
        joined = " ".join(c for c in cells if c).lower()
        if "size specifications" in joined or (
            any(c.lower() == "size ref." for c in cells) and any("style" in c.lower() for c in cells)
        ):
            header_idx = i
            break
    if header_idx is None:
        return None

    header = rows[header_idx]
    size_col = None
    size_label = None
    for j, cell in enumerate(header):
        label = clean(cell)
        low = label.lower()
        if not low:
            continue
        if low in {"size specifications", "size ref.", "tole -/+", "tole  -/+", "grading", "model :", "model:"}:
            continue
        if re.fullmatch(r"[mls]|xl|xxl|xxxl|xxxxl|\d?xl|[0-9]{1,2}", low) or re.fullmatch(
            r"[0-9]+\s*cm", low
        ):
            size_col = j
            size_label = label
            break
    # Fallback: first column after col 7 that has numeric values in data rows
    if size_col is None:
        for j in range(7, max(len(r or []) for r in rows[header_idx + 1 : header_idx + 20] or [[]]) or 14):
            nums = 0
            for row in rows[header_idx + 1 : header_idx + 25]:
                if j < len(row or []) and to_number((row or [])[j]) is not None:
                    nums += 1
            if nums >= 4:
                size_col = j
                size_label = clean((header[j] if j < len(header) else None) or f"col{j}")
                break
    if size_col is None:
        return None

    # Customer / fabric from rows above
    name = None
    fabric_code = None
    pattern_ref = None
    for row in rows[: header_idx + 1]:
        cells = [clean(c) for c in (row or [])]
        for j, cell in enumerate(cells):
            low = cell.lower().rstrip(":")
            if low.startswith("customer") and j + 1 < len(cells) and cells[j + 1]:
                name = cells[j + 1]
            if low in {"fabric", "fabric code", "fabric:"} and j + 1 < len(cells):
                fabric_code = cells[j + 1] or fabric_code
            if cell.startswith("MD-SHORT") or cell.startswith("MD-"):
                pattern_ref = cell

    points = []
    for row in rows[header_idx + 1 :]:
        cells = row or []
        # Point name typically column index 4
        point_name = ""
        for idx in (4, 0, 1, 3):
            if idx < len(cells):
                candidate = clean(cells[idx])
                if candidate and not re.fullmatch(r"[A-Z]{1,3}\d?", candidate) and to_number(candidate) is None:
                    # skip pure letter refs like A, C, BB
                    if idx == 4 or (idx == 0 and "waist" in candidate.lower()):
                        point_name = candidate
                        break
        if not point_name:
            # col 4 preferred even if letter-like skipped above failed
            point_name = clean(cells[4]) if len(cells) > 4 else ""
        if not point_name:
            continue
        if point_name.lower().startswith("special"):
            break
        # Skip letter codes mistaken as names
        if re.fullmatch(r"[A-Z]{1,3}\d?", point_name):
            continue
        size_val = to_number(cells[size_col]) if size_col < len(cells) else None
        if size_val is None:
            continue
        points.append(
            {
                "point_id": slugify(point_name),
                "name": point_name,
                "base_value": size_val,
                "target_value": size_val,
                "trial_values": {},
                "final_value": None,
                "remarks": None,
            }
        )

    if not points:
        return None

    # MD-SHORT factory sheets are almost always cm (half-waist often 40-60).
    waist = next(
        (
            p["base_value"]
            for p in points
            if p["base_value"] is not None and "waist" in p["name"].lower() and "band" not in p["name"].lower()
        ),
        None,
    )
    if waist is not None:
        unit = "cm" if waist >= 30 else "in"
    else:
        vals = [p["base_value"] for p in points if p["base_value"] is not None]
        med = sorted(vals)[len(vals) // 2] if vals else 0
        unit = "cm" if med >= 30 else "in"

    return {
        "ok": True,
        "path": str(path),
        "filename": path.name,
        "pattern_ref": pattern_ref,
        "client_name": name,
        "description": "MD-SHORT factory sheet",
        "fabric_code": fabric_code,
        "sheet_stage": "final",
        "size_label": size_label,
        "unit": unit,
        "order_date": None,
        "special_instructions": None,
        "points": points,
        "filled_count": sum(1 for p in points if p["base_value"] is not None),
        "format": "md-short",
    }


def parse_sheet(path: Path) -> dict:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    header_idx = None
    for i, row in enumerate(rows):
        first = clean(row[0] if row else "")
        if first.lower().startswith("measurement point"):
            header_idx = i
            break
    if header_idx is None:
        md = parse_md_short_sheet(path, rows)
        if md:
            return md
        return {"ok": False, "error": "no measurement header", "path": str(path)}

    header = rows[header_idx]
    size_candidates: list[tuple[int, str, int]] = []  # (col, label, priority)
    trial_cols: dict[int, int] = {}
    final_col = None
    remarks_col = None
    for j, cell in enumerate(header):
        label = clean(cell)
        low = label.lower()
        if not low:
            continue
        if low == "remarks":
            remarks_col = j
        elif low == "final" or low.startswith("final"):
            final_col = j
        elif "trial" in low:
            m = re.search(r"(\d+)", low)
            trial_cols[int(m.group(1)) if m else 1] = j
        elif low in {"sewing", "sewimg", "adjustment"}:
            continue
        elif (
            "pattern size" in low
            or low == "client body"
            or low.startswith("size")
            or low.startswith("r-")
            or low.startswith("l-")
            or low.startswith("s-")
            or re.fullmatch(r"[mls]|xl|xxl|xxxl|\d{2}", low)
            # House size labels like "GL- 54", "FR- 48"
            or re.fullmatch(r"(gl|fr|fd|ju)\s*[-_]?\s*\d{1,3}", low)
        ):
            # Prefer pattern size / size / house size over client body when both have values
            if "pattern size" in low or low.startswith("size"):
                priority = 3
            elif re.fullmatch(r"(gl|fr|fd|ju)\s*[-_]?\s*\d{1,3}", low):
                priority = 2
            elif low == "client body":
                priority = 1
            else:
                priority = 2
            size_candidates.append((j, label, priority))

    size_col = None
    size_label = None
    if size_candidates:
        # Score by how many numeric values exist under each candidate column
        scored = []
        for j, label, priority in size_candidates:
            nums = 0
            for row in rows[header_idx + 1 : header_idx + 40]:
                cells = row or []
                if j < len(cells) and to_number(cells[j]) is not None:
                    nums += 1
            scored.append((nums, priority, j, label))
        scored.sort(key=lambda t: (t[0], t[1]), reverse=True)
        if scored and scored[0][0] > 0:
            size_col = scored[0][2]
            size_label = scored[0][3]
        else:
            # Fall back to highest-priority header even if empty
            size_candidates.sort(key=lambda t: t[2], reverse=True)
            size_col = size_candidates[0][0]
            size_label = size_candidates[0][1]

    # Metadata above header
    pattern_ref = None
    description = None
    fabric_code = None
    sheet_stage = None
    name = None
    order_date = None
    for row in rows[:header_idx]:
        cells = [clean(c) for c in (row or [])]
        joined = " | ".join(c for c in cells if c)
        low_join = joined.lower()
        for j, cell in enumerate(cells):
            low = cell.lower().rstrip(":")
            if "pattern" in low and "ref" in low:
                pattern_ref = clean(
                    next((c for c in cells[j + 1 :] if c and c != ":"), "")
                ) or pattern_ref
            if low == "name":
                name = clean(next((c for c in cells[j + 1 :] if c and c != ":"), "")) or name
            if low == "description":
                description = (
                    clean(next((c for c in cells[j + 1 :] if c and c != ":"), "")) or description
                )
            if "fabric" in low and ("name" in low or "code" in low or low.endswith(":")):
                fabric_code = (
                    clean(next((c for c in cells[j + 1 :] if c and c != ":"), "")) or fabric_code
                )
            if low in {"order date", "orderdate"}:
                order_date = (
                    clean(next((c for c in cells[j + 1 :] if c and c != ":"), "")) or order_date
                )
        if re.search(r"\bfinal\b", low_join) and "measurement" not in low_join:
            sheet_stage = sheet_stage or "final"
        if re.search(r"\btrial\b", low_join) and "measurement" not in low_join:
            sheet_stage = sheet_stage or "trial"

    unit = "cm" if size_label and "cm" in size_label.lower() else None
    points = []
    special = None
    for row in rows[header_idx + 1 :]:
        cells = row or []
        first = clean(cells[0] if cells else "")
        if not first:
            continue
        if first.lower().startswith("special instruction"):
            trailing = clean(first.split(":", 1)[1]) if ":" in first else ""
            rest = " ".join(clean(c) for c in cells[1:] if clean(c))
            special = clean(f"{trailing} {rest}") or None
            break
        size_val = to_number(cells[size_col]) if size_col is not None and size_col < len(cells) else None
        trial_values = {}
        for n, col in trial_cols.items():
            val = to_number(cells[col]) if col < len(cells) else None
            if val is not None:
                trial_values[str(n)] = val
        final_val = (
            to_number(cells[final_col]) if final_col is not None and final_col < len(cells) else None
        )
        remarks = (
            clean(cells[remarks_col])
            if remarks_col is not None and remarks_col < len(cells)
            else ""
        )
        if size_val is None and not trial_values and final_val is None and not remarks:
            continue
        points.append(
            {
                "point_id": slugify(first),
                "name": first,
                "base_value": size_val,
                "target_value": size_val,
                "trial_values": trial_values,
                "final_value": final_val,
                "remarks": remarks or None,
            }
        )

    # Infer unit: prefer Total Length magnitude, else median. Length > 50 => cm.
    if unit is None and points:
        length = next(
            (
                p["base_value"]
                for p in points
                if p["base_value"] is not None
                and "length" in p["name"].lower()
                and "slv" not in p["name"].lower()
                and "sleeve" not in p["name"].lower()
            ),
            None,
        )
        if length is not None:
            unit = "cm" if length >= 50 else "in"
        else:
            vals = [p["base_value"] for p in points if p["base_value"] is not None]
            if vals:
                med = sorted(vals)[len(vals) // 2]
                # Shorts waist-extend in cm is often 45-60; half-waist inches is < 30.
                unit = "cm" if med >= 35 else "in"

    return {
        "ok": True,
        "path": str(path),
        "filename": path.name,
        "pattern_ref": pattern_ref,
        "client_name": name,
        "description": description,
        "fabric_code": fabric_code,
        "sheet_stage": sheet_stage,
        "size_label": size_label,
        "unit": unit or "in",
        "order_date": order_date,
        "special_instructions": special,
        "points": points,
        "filled_count": sum(1 for p in points if p["base_value"] is not None),
    }


def main():
    paths = [Path(p) for p in sys.argv[1:]]
    out = [parse_sheet(p) for p in paths]
    json.dump(out if len(out) != 1 else out[0], sys.stdout, indent=2)
    sys.stdout.write("\n")


if __name__ == "__main__":
    main()
