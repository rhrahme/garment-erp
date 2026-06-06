#!/usr/bin/env python3
"""Import Stylbiella product info Excel into supplier catalog JSON."""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("Install openpyxl: pip3 install openpyxl") from exc

FABRIC_CODE_RE = re.compile(r"^\d{4,6}/\d{2,4}$")


def clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def parse_features(raw) -> dict:
    text = clean(raw)
    if not text:
        return {"composition": None, "weight_gsm": None, "width_cm": None, "quality": None}

    weight_match = re.search(r"(\d+)\s*GR\b", text, re.I)
    width_match = re.search(r"(\d+)\s*CM\b", text, re.I)
    composition_part = re.sub(r"-?\d+\s*GR\b.*$", "", text, flags=re.I)
    composition_part = re.sub(r"-?\d+\s*CM\b.*$", "", composition_part, flags=re.I).strip()

    quality = None
    quality_match = re.match(
        r"^(SUPER\d+'?S?|HIGH TWIST|HIGHT TWIST|TROPICAL|SOLBI\s*AIR|CASHCO|CROSS[- ]PLY)",
        composition_part,
        re.I,
    )
    if quality_match:
        quality = re.sub(r"\s+", " ", quality_match.group(1)).strip()
        composition_part = composition_part[quality_match.end() :].lstrip("- ").strip()

    composition_part = re.sub(r"(\d+%)([A-Za-z])", r"\1 \2", composition_part)
    composition_part = re.sub(r"([A-Za-z])(\d+%)", r"\1 \2", composition_part)
    composition_part = re.sub(r"\s+", " ", composition_part).strip()

    return {
        "composition": composition_part or quality or text,
        "weight_gsm": int(weight_match.group(1)) if weight_match else None,
        "width_cm": int(width_match.group(1)) if width_match else None,
        "quality": quality,
    }


def build_description(**parts) -> str:
    return " — ".join(value for value in parts.values() if value)


def find_lookbook_columns(ws) -> dict | None:
    for row_idx in range(3, 8):
        row = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))
        labels = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if not any("product features" in label for label in labels):
            continue

        def col(label: str, occurrence: int = 0) -> int | None:
            hits = [index for index, value in enumerate(labels) if label in value]
            return hits[occurrence] if len(hits) > occurrence else None

        return {
            "book_number": 0,
            "book_name": 1,
            "garments": 3,
            "suit_features": col("product features", 0),
            "suit_code": col("product code", 0),
            "suit_comment": col("comments", 0),
            "suit_pattern": col("pattern", 0),
            "suit_color": col("color", 0),
            "shirt_code": col("product code", 1),
            "shirt_features": col("product features", 1),
            "data_start_row": row_idx + 1,
        }
    return None


def import_lookbooks(ws, add) -> None:
    columns = find_lookbook_columns(ws)
    if not columns:
        return

    for row in ws.iter_rows(min_row=columns["data_start_row"], values_only=True):
        book_number = row[columns["book_number"]] if len(row) > columns["book_number"] else None
        book_name = clean(row[columns["book_name"]] if len(row) > columns["book_name"] else None)
        garments = clean(row[columns["garments"]] if len(row) > columns["garments"] else None)

        suit_code = (
            clean(row[columns["suit_code"]]) if columns["suit_code"] is not None and len(row) > columns["suit_code"] else None
        )
        suit_features = (
            row[columns["suit_features"]]
            if columns["suit_features"] is not None and len(row) > columns["suit_features"]
            else None
        )
        suit_comment = (
            clean(row[columns["suit_comment"]])
            if columns["suit_comment"] is not None and len(row) > columns["suit_comment"]
            else None
        )
        suit_pattern = (
            clean(row[columns["suit_pattern"]])
            if columns["suit_pattern"] is not None and len(row) > columns["suit_pattern"]
            else None
        )
        suit_color = (
            clean(row[columns["suit_color"]])
            if columns["suit_color"] is not None and len(row) > columns["suit_color"]
            else None
        )
        shirt_code = (
            clean(row[columns["shirt_code"]])
            if columns["shirt_code"] is not None and len(row) > columns["shirt_code"]
            else None
        )
        shirt_features = (
            row[columns["shirt_features"]]
            if columns["shirt_features"] is not None and len(row) > columns["shirt_features"]
            else None
        )

        if suit_code:
            parsed = parse_features(suit_features)
            add(
                {
                    "fabric_number": suit_code,
                    "book_number": str(book_number) if book_number is not None else None,
                    "collection": book_name,
                    "composition": parsed["composition"],
                    "color": suit_color,
                    "description": build_description(
                        book=book_name,
                        category=garments,
                        quality=parsed["quality"],
                        pattern=suit_pattern,
                        color=suit_color,
                        comment=suit_comment,
                    ),
                    "weight_gsm": parsed["weight_gsm"],
                    "width_cm": parsed["width_cm"],
                    "unit_price": None,
                    "unit": "meters",
                    "currency": "USD",
                    "is_active": True,
                    "category": (garments or "suiting").lower(),
                }
            )

        if shirt_code:
            parsed = parse_features(shirt_features)
            add(
                {
                    "fabric_number": shirt_code,
                    "book_number": str(book_number) if book_number is not None else None,
                    "collection": book_name,
                    "composition": parsed["composition"],
                    "color": None,
                    "description": build_description(
                        book=book_name,
                        category="SHIRTS",
                        quality=parsed["quality"],
                    ),
                    "weight_gsm": parsed["weight_gsm"],
                    "width_cm": parsed["width_cm"],
                    "unit_price": None,
                    "unit": "meters",
                    "currency": "USD",
                    "is_active": True,
                    "category": "shirts",
                }
            )


def import_standard_bunches(ws, add) -> None:
    for row in ws.iter_rows(min_row=3, values_only=True):
        bunch_number, _, bunch_name, code, features, comment, pattern = (list(row) + [None] * 7)[:7]
        code = clean(code)
        if not code:
            continue
        parsed = parse_features(features)
        add(
            {
                "fabric_number": code,
                "book_number": str(bunch_number) if bunch_number is not None else None,
                "collection": clean(bunch_name),
                "composition": parsed["composition"],
                "color": None,
                "description": build_description(
                    book=clean(bunch_name),
                    category="BUNCH",
                    quality=parsed["quality"],
                    pattern=clean(pattern),
                    comment=clean(comment),
                ),
                "weight_gsm": parsed["weight_gsm"],
                "width_cm": parsed["width_cm"],
                "unit_price": None,
                "unit": "meters",
                "currency": "USD",
                "is_active": True,
                "category": "bunch",
            }
        )


def import_charm_bunch_grid(ws, add, bunch_name: str) -> None:
    for row in ws.iter_rows(min_row=4, values_only=True):
        for cell in row:
            code = clean(cell)
            if code and FABRIC_CODE_RE.match(code):
                add(
                    {
                        "fabric_number": code,
                        "book_number": None,
                        "collection": bunch_name,
                        "composition": None,
                        "color": None,
                        "description": build_description(book=bunch_name, category="BUNCH"),
                        "weight_gsm": None,
                        "width_cm": None,
                        "unit_price": None,
                        "unit": "meters",
                        "currency": "USD",
                        "is_active": True,
                        "category": "bunch",
                    }
                )


def is_charm_grid_sheet(ws) -> bool:
    rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
    if len(rows) < 2:
        return False
    header = [str(cell).strip().lower() if cell is not None else "" for cell in rows[1]]
    return header.count("product code") >= 2


def import_workbook(xlsx_path: Path) -> list[dict]:
    fabrics: list[dict] = []
    seen: set[str] = set()

    def add(entry: dict) -> None:
        fabric_number = entry["fabric_number"]
        if not fabric_number or fabric_number in seen:
            return
        seen.add(fabric_number)
        fabrics.append(entry)

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if "Lookbooks" in workbook.sheetnames:
        import_lookbooks(workbook["Lookbooks"], add)

    bunch_sheets = [name for name in workbook.sheetnames if name == "Bunches" or name.startswith("Bunch-")]
    for sheet_name in bunch_sheets:
        ws = workbook[sheet_name]
        if is_charm_grid_sheet(ws):
            import_charm_bunch_grid(ws, add, sheet_name.replace("Bunch-", "").strip())
        else:
            import_standard_bunches(ws, add)

    fabrics.sort(key=lambda fabric: fabric["fabric_number"])
    return fabrics


def main() -> None:
    parser = argparse.ArgumentParser(description="Import Stylbiella product info Excel")
    parser.add_argument("xlsx", nargs="?", help="Path to Stylbiella product info .xlsx")
    parser.add_argument("--season", default="SS26", help="Season label, e.g. AW25, SS25, SS26")
    parser.add_argument(
        "--out",
        default=None,
        help="Output JSON path (default: src/data/suppliers/stylbiella-{season}.json)",
    )
    args = parser.parse_args()

    season = args.season.strip().lower()
    xlsx_path = Path(args.xlsx) if args.xlsx else Path.home() / "Downloads/Stylbiella SS26 Product info.xlsx"
    out_path = Path(args.out) if args.out else Path(f"src/data/suppliers/stylbiella-{season}.json")

    if not xlsx_path.exists():
        raise SystemExit(f"File not found: {xlsx_path}")

    fabrics = import_workbook(xlsx_path)
    payload = {
        "document_type": "price_list",
        "supplier": {
            "code": "STYLBIELLA",
            "name": "Stylbiella",
            "country": "Italy",
            "is_fabric_supplier": True,
            "lead_time_days": 14,
            "currency": "USD",
        },
        "price_list_name": f"{args.season.upper()} Product Info",
        "imported_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source_file": xlsx_path.name,
        "fabric_count": len(fabrics),
        "fabrics": fabrics,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2) + "\n")
    print(f"✓ Stylbiella {args.season.upper()}: {len(fabrics)} fabrics → {out_path}")


if __name__ == "__main__":
    main()
