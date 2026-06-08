#!/usr/bin/env python3
"""Import Canclini linen warehouse stock from Linen Stock HAGAN.xlsx."""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    raise

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Desktop" / "Fabrics" / "Linen Stock HAGAN.xlsx"
OUT_PATH = ROOT / "src" / "data" / "suppliers" / "canclini-linen-stock.json"


def parse_width(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"(\d+)", str(value))
    return int(match.group(1)) if match else None


def parse_weight(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[\d.]+", str(value))
    return float(match.group(0)) if match else None


def is_valid_code(code: str | None) -> bool:
    if not code:
        return False
    text = str(code).strip()
    if not text or text.lower().startswith("total"):
        return False
    return bool(re.match(r"^\d{2}[TH]\d{3}$", text, re.IGNORECASE))


def weight_category(code: str, weight_gsm: float | None) -> str:
    if "H" in code.upper()[2:4]:
        return "H-weight"
    return "T-weight"


def parse_sheet(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Codes"]
    fabrics: list[dict] = []
    seen: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        pairs = [
            (row[0], row[1], row[2], row[3], row[4]),
            (row[5], row[6], row[7], row[8], row[9]),
        ]
        for code, composition, weight, width, avail_meters in pairs:
            if not is_valid_code(code):
                continue
            fabric_number = str(code).strip().upper()
            if fabric_number in seen:
                continue
            seen.add(fabric_number)

            weight_gsm = parse_weight(weight)
            width_cm = parse_width(width)
            category = weight_category(fabric_number, weight_gsm)
            weight_label = f"{int(weight_gsm)} g/m²" if weight_gsm else "unknown weight"

            fabrics.append(
                {
                    "fabric_number": fabric_number,
                    "composition": str(composition).strip() if composition else "100% Linen",
                    "color": None,
                    "description": f"Canclini linen — {category} ({weight_label})",
                    "weight_gsm": weight_gsm,
                    "width_cm": width_cm,
                    "collection": "HAGAN Linen Stock",
                    "category": category,
                    "unit_price": None,
                    "unit": "meters",
                    "currency": "EUR",
                    "is_active": True,
                    "stock_status": "in_stock",
                    "available_meters": float(avail_meters) if isinstance(avail_meters, (int, float)) else None,
                }
            )

    wb.close()
    fabrics.sort(key=lambda item: item["fabric_number"])
    return fabrics


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    fabrics = parse_sheet(xlsx_path)
    payload = {
        "document_type": "warehouse_stock",
        "supplier": {
            "code": "CANCLINI",
            "name": "Canclini",
            "country": "Italy",
            "is_fabric_supplier": True,
            "lead_time_days": 0,
            "currency": "EUR",
        },
        "price_list_name": "HAGAN Linen Stock",
        "imported_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "source_file": xlsx_path.name,
        "fabric_count": len(fabrics),
        "fabrics": fabrics,
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(f"Canclini linen stock: {len(fabrics)} fabrics -> {OUT_PATH}")


if __name__ == "__main__":
    main()
